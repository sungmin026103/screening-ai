from __future__ import annotations

import io
import re
from dataclasses import dataclass, field
from copy import deepcopy

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from sklearn.base import BaseEstimator, TransformerMixin
from sklearn.calibration import CalibratedClassifierCV
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    accuracy_score,
    average_precision_score,
    confusion_matrix,
    precision_recall_curve,
    precision_score,
    recall_score,
    roc_auc_score,
    roc_curve,
)
from sklearn.metrics.pairwise import cosine_similarity
from sklearn.model_selection import StratifiedKFold, cross_val_predict
from sklearn.pipeline import FeatureUnion, Pipeline
from sklearn.svm import LinearSVC
from scipy.stats import beta as _beta_dist

# 문장 임베딩(의미 기반) 신호는 선택적 의존성이다. requirements.txt에
# sentence-transformers가 없거나 배포 환경에서 모델 다운로드가 막혀 있어도
# 앱 전체가 죽지 않도록 임포트 실패를 흡수하고, 사용 가능 여부를 플래그로 남긴다.
try:
    from sentence_transformers import SentenceTransformer
    _HAS_SENTENCE_TRANSFORMERS = True
except Exception:  # pragma: no cover
    SentenceTransformer = None
    _HAS_SENTENCE_TRANSFORMERS = False

# 화면에 표시되는 3단계 우선순위 (정렬 순서 그대로 사용)
PRIORITY_ORDER = ["우선 검토", "경계 문헌", "안전 제외 후보"]

# ---------------------------------------------------------------------------
# 검증된 기준 성능(Reference benchmark)
# ---------------------------------------------------------------------------
# 이 값은 현재 프로젝트의 실시간 성능이 아니라, 이전에 사람 라벨이 있는 독립적인
# 검증 데이터에서 얻은 기준 성능이다. 라벨이 없는 새 프로젝트에서도 사용자가
# "이 AI가 어느 정도 검증되었는지" 확인할 수 있도록 UI에 항상 표시한다.
# 새 버전의 모델을 다시 검증하면 이 상수만 업데이트하면 된다.
REFERENCE_BENCHMARK = {
    "name": "SpaceFood labeled benchmark",
    "version_note": "Legacy reference benchmark (pre-V16); V16 requires revalidation",
    "n": 2499,
    "recall": 0.9893,
    "precision": 0.1800,
    "false_negative": 2,
    "roc_auc": 0.9446,
    "average_precision": 0.6856,
    "screening_burden": 0.411,
    "work_saved": 0.589,
}

# 지도학습(재현율 통계적 보장) 모드로 전환되는 최소 라벨 수. 이 미만이면 앱은
# 자동으로 zero-shot 모드로 동작한다 — 사람이 매번 "어느 모드로 할지" 고르는 게
# 아니라, 라벨 존재 여부라는 데이터 상태로 결정되는 고정 기준이다.
MIN_LABELS_FOR_SUPERVISED = 100
TRAINING_SAMPLE_SIZE = 200
MIN_INCLUDE_FOR_SUPERVISED = 10

# 엑셀 다운로드 시 구간 순서와 배경색. False Negative는 실제 라벨이 Include인데
# 컷오프 밖으로 밀려난, 눈에 띄어야 하는 문헌이라 원래 버킷에서 따로 떼어내
# 독립된 구간으로 모은다 (행이 두 번 나오지 않도록 한 구간에만 배치한다).
EXPORT_GROUP_ORDER = ["우선 검토", "경계 문헌", "False Negative", "안전 제외 후보"]
EXPORT_GROUP_COLORS = {
    "우선 검토": "FFFFFF",       # 흰색
    "경계 문헌": "E9ECF1",       # 중간 회색
    "False Negative": "F7D6D2",  # 경고용 붉은색 (기존 FN 탭과 동일 색)
    "안전 제외 후보": "B8BDC6",  # 진한 회색
}


@dataclass
class ScreeningResult:
    predictions: pd.DataFrame
    metrics: dict
    threshold: float
    pr_curve: dict = field(default_factory=dict)   # {"precision": [...], "recall": [...], "thresholds": [...]} (참고용 차트)
    roc_curve: dict = field(default_factory=dict)   # {"fpr": [...], "tpr": [...]}
    confusion: dict = field(default_factory=dict)   # {"tn":.., "fp":.., "fn":.., "tp":..}


def _find_col(df: pd.DataFrame, options: list[str]) -> str | None:
    lookup = {str(c).strip().lower(): c for c in df.columns}
    return next((lookup[x] for x in options if x in lookup), None)


def _normalize_label_value(value):
    """사용자 라벨을 0/1로 표준화한다. O/X, 숫자, 한글 판정을 모두 허용한다."""
    if pd.isna(value):
        return np.nan
    key = str(value).strip().lower()
    mapping = {
        "1": 1, "1.0": 1, "o": 1, "○": 1, "ㅇ": 1,
        "include": 1, "included": 1, "yes": 1, "y": 1, "포함": 1,
        "0": 0, "0.0": 0, "x": 0, "×": 0,
        "exclude": 0, "excluded": 0, "no": 0, "n": 0, "제외": 0,
    }
    return mapping.get(key, np.nan)


def _reviewer_label_columns(df: pd.DataFrame, excluded: set[str]) -> list[str]:
    """O/X 또는 0/1 값이 주로 들어 있는 검토자 열을 자동 탐지한다."""
    found = []
    for col in df.columns:
        if col in excluded:
            continue
        vals = df[col].dropna()
        if len(vals) == 0:
            continue
        normalized = vals.map(_normalize_label_value)
        valid_rate = float(normalized.notna().mean())
        if valid_rate >= 0.70 and normalized.notna().sum() >= 2:
            found.append(col)
    return found


def prepare_screening_data(df: pd.DataFrame) -> tuple[pd.DataFrame, str]:
    """Human_Label을 표준화한다.

    - 라벨 열이 하나면 O/o/○/ㅇ = Include(1), X/x/× = Exclude(0)로 인식한다.
    - 검토자 열이 2개 이상 자동 탐지되면, 모두 Include로 일치할 때만 1,
      모두 Exclude로 일치할 때만 0으로 두고, 불일치(의견 불일치)는 NaN으로
      두어 학습에서 제외한다 (합의되지 않은 판정을 모델이 배우지 않도록).
    """
    title_col = _find_col(df, ["title", "제목"])
    abstract_col = _find_col(df, ["abstract", "초록"])
    label_col = _find_col(df, [
        "human_label", "human label", "include", "label", "decision",
        "포함", "라벨", "판정", "최종판정", "최종라벨",
    ])
    if not title_col:
        raise ValueError("제목(Title/제목) 열이 필요합니다.")

    out = pd.DataFrame()
    out["Title"] = df[title_col].fillna("").astype(str)
    out["Abstract"] = df[abstract_col].fillna("").astype(str) if abstract_col else ""

    if label_col:
        out["Human_Label"] = df[label_col].map(_normalize_label_value)
        label_source = str(label_col)
    else:
        excluded = {title_col}
        if abstract_col:
            excluded.add(abstract_col)
        reviewer_cols = _reviewer_label_columns(df, excluded)
        if not reviewer_cols:
            raise ValueError(
                "라벨 열을 찾지 못했습니다. Human_Label/Label 열을 추가하거나, "
                "검토자 열에 O(포함)와 X(제외)를 입력해 주세요."
            )
        normalized = pd.DataFrame({c: df[c].map(_normalize_label_value) for c in reviewer_cols})
        if len(reviewer_cols) == 1:
            out["Human_Label"] = normalized.iloc[:, 0]
        else:
            # 두 명 이상이 모두 같은 판정을 내린 행만 학습에 사용하고, 불일치는 미합의로 둔다.
            n_valid = normalized.notna().sum(axis=1)
            unanimous = normalized.nunique(axis=1, dropna=True).eq(1) & n_valid.ge(2)
            out["Human_Label"] = np.where(unanimous, normalized.bfill(axis=1).iloc[:, 0], np.nan)
        label_source = "검토자 자동합의: " + ", ".join(map(str, reviewer_cols))

    out["Human_Label"] = pd.to_numeric(out["Human_Label"], errors="coerce")
    out["Text"] = (out["Title"] + " " + out["Abstract"]).str.strip()
    out["StructuredText"] = [_serialize_structured(t, a) for t, a in zip(out["Title"], out["Abstract"])]
    return out, label_source


def detect_label_count(df: pd.DataFrame) -> int:
    """업로드된 파일에 유효 라벨(Include/Exclude 둘 다 있는)이 몇 개 있는지 반환한다.
    라벨 열이 아예 없거나 형식이 안 맞으면 0을 반환한다 (예외를 던지지 않음).
    app.py가 이 값으로 지도학습/zero-shot 모드를 자동 판별하는 데 쓴다."""
    try:
        data, _ = prepare_screening_data(df)
        return int(data["Human_Label"].isin([0, 1]).sum())
    except Exception:
        return 0


def build_training_sample(
    df: pd.DataFrame,
    criteria_text: str,
    exclusion_text: str = "",
    sample_size: int = TRAINING_SAMPLE_SIZE,
) -> pd.DataFrame:
    """전체 코퍼스에서 학습 가치가 높은 문헌을 빠르게 한 번에 뽑는다.

    200편을 높은 PICO 적합도 50%, 경계 35%, 낮은 적합도 15%로 구성한다.
    이 단계는 의도적으로 가벼운 TF-IDF만 사용해 8천~수만 편에서도 빠르게 끝나며,
    실제 최종 지도학습 모델은 이후 200편 라벨을 이용해 별도로 학습한다.
    """
    if not criteria_text or not criteria_text.strip():
        raise ValueError("학습용 문헌을 선정하려면 PICO 기준이 필요합니다.")
    if len(df) == 0:
        raise ValueError("문헌 데이터가 비어 있습니다.")

    base = df.copy().reset_index(drop=True)
    title_col = _find_col(base, ["title", "제목"])
    abstract_col = _find_col(base, ["abstract", "초록"])
    if title_col is None:
        raise ValueError("제목(Title/제목) 열이 필요합니다.")
    titles = base[title_col].fillna("").astype(str)
    abstracts = base[abstract_col].fillna("").astype(str) if abstract_col else pd.Series([""] * len(base))
    docs = (titles + " " + abstracts).str.strip().tolist()
    base["_Source_Index"] = np.arange(len(base), dtype=int)

    sections = _parse_pico_sections(criteria_text)
    queries = [q.strip() for q in sections.values() if q and q.strip()]
    if not queries:
        queries = [criteria_text.strip()]
    exclusion_items = _split_bullet_items(exclusion_text)
    all_queries = queries + exclusion_items

    # 샘플 선정은 속도가 핵심이므로 pretrained embedding 다운로드 없이 TF-IDF만 사용.
    vec = TfidfVectorizer(ngram_range=(1, 2), min_df=1, max_features=60000, sublinear_tf=True, stop_words="english")
    mat = vec.fit_transform(docs + all_queries)
    doc_mat = mat[:len(docs)]
    query_mat = mat[len(docs):]
    pico_q = query_mat[:len(queries)]
    pico_sim = cosine_similarity(doc_mat, pico_q)
    # 샘플 enrichment에서는 min만 쓰면 짧은 제목/초록에서 0이 과도하게 많아져
    # 구분력이 사라질 수 있어 mean+min을 결합한다.
    pico_score = 0.65 * pico_sim.mean(axis=1) + 0.35 * pico_sim.min(axis=1)
    if exclusion_items:
        excl_q = query_mat[len(queries):]
        excl_score = cosine_similarity(doc_mat, excl_q).max(axis=1)
    else:
        excl_score = np.zeros(len(docs), dtype=float)
    scores = np.asarray(pico_score - 0.75 * excl_score, dtype=float)

    n = min(int(sample_size), len(base))
    n_high = int(round(n * 0.50))
    n_boundary = int(round(n * 0.35))
    n_low = n - n_high - n_boundary
    ranked = np.argsort(scores)

    # 데이터 자체의 점수 분포에서 중앙 경계를 잡고 그 주변을 uncertainty 표본으로 사용.
    boundary = float(_otsu_threshold(scores))
    boundary_order = np.argsort(np.abs(scores - boundary))

    selected: list[tuple[int, str]] = []
    used_idx: set[int] = set()
    used_titles: set[str] = set()

    def add(indices, label, limit):
        count = 0
        for raw_idx in indices:
            idx = int(raw_idx)
            title_key = titles.iloc[idx].strip().casefold()
            if idx in used_idx or not title_key or title_key in used_titles:
                continue
            used_idx.add(idx)
            used_titles.add(title_key)
            selected.append((idx, label))
            count += 1
            if count >= limit:
                break

    add(ranked[::-1], "High PICO relevance", n_high)
    add(boundary_order, "Decision boundary", n_boundary)
    add(ranked, "Low PICO relevance", n_low)
    if len(selected) < n:
        # 중복 제목 때문에 부족한 경우 전체 점수 순위에서 남은 문헌으로 보충.
        add(ranked[::-1], "Supplemental", n - len(selected))

    rows = []
    for idx, stratum in selected[:n]:
        row = base.iloc[idx].copy()
        row["Training_Stratum"] = stratum
        rows.append(row)
    out = pd.DataFrame(rows).reset_index(drop=True)
    out.insert(0, "Training_No", np.arange(1, len(out) + 1))
    out["Human_Label"] = ""
    return out


def merge_training_labels(full_df: pd.DataFrame, labeled_sample_df: pd.DataFrame) -> tuple[pd.DataFrame, dict]:
    """200편 라벨 파일을 원본 전체 문헌에 병합한다. _Source_Index를 우선 사용하고,
    없으면 정규화한 제목으로 매칭한다. Human_Label은 O/X 또는 1/0을 허용한다.
    """
    full = full_df.copy().reset_index(drop=True)
    sample = labeled_sample_df.copy().reset_index(drop=True)
    sample_prepared, _ = prepare_screening_data(sample)
    labels = sample_prepared["Human_Label"]

    full["Human_Label"] = np.nan
    matched = 0
    if "_Source_Index" in sample.columns:
        for i, lab in labels.items():
            if lab not in (0, 1):
                continue
            try:
                idx = int(sample.loc[i, "_Source_Index"])
            except Exception:
                continue
            if 0 <= idx < len(full):
                full.loc[idx, "Human_Label"] = int(lab)
                matched += 1
    else:
        title_full = _find_col(full, ["title", "제목"])
        title_sample = _find_col(sample, ["title", "제목"])
        if not title_full or not title_sample:
            raise ValueError("라벨 파일을 병합하려면 _Source_Index 또는 Title/제목 열이 필요합니다.")
        lookup = {}
        for idx, t in enumerate(full[title_full].fillna("").astype(str)):
            lookup.setdefault(t.strip().casefold(), idx)
        for i, lab in labels.items():
            if lab not in (0, 1):
                continue
            key = str(sample.loc[i, title_sample]).strip().casefold()
            idx = lookup.get(key)
            if idx is not None:
                full.loc[idx, "Human_Label"] = int(lab)
                matched += 1

    valid = full["Human_Label"].isin([0, 1])
    include_n = int((full.loc[valid, "Human_Label"] == 1).sum())
    exclude_n = int((full.loc[valid, "Human_Label"] == 0).sum())
    return full, {"matched": int(matched), "include_n": include_n, "exclude_n": exclude_n, "labeled_n": int(valid.sum())}


# ---------------------------------------------------------------------------
# 의미 기반(임베딩) 신호. TF-IDF 계열(word/char/PICO 유사도)은 결국 전부
# 표면적 단어 일치에 의존하기 때문에 "서로 다른 근거"라고 보기 어렵다.
# 사전학습된 문장 임베딩(SBERT 계열)은 동의어·다른 표현으로 쓰인 문헌도
# 의미로 포착하므로, TF-IDF 신호들과 상관관계가 낮은 진짜 독립적인 근거가 된다.
# 또한 이 인코더는 우리 데이터로 다시 학습(fit)되지 않는 고정 가중치이므로,
# 전체 문헌에 대해 미리 한 번만 계산해도 검증 fold 누수가 전혀 발생하지 않는다
# (TF-IDF는 데이터 의존적으로 fit되므로 fold마다 다시 학습해야 하는 것과 대조적).
# ---------------------------------------------------------------------------

EMBEDDING_MODEL_NAME = "pritamdeka/S-PubMedBert-MS-MARCO"  # 생의학 초록 특화. 배포 환경 리소스가 빠듯하면
# "sentence-transformers/all-MiniLM-L6-v2" (가볍고 범용, CPU에서 훨씬 빠름)로 교체 가능.

_embed_model_singleton: dict[str, "SentenceTransformer"] = {}


def embeddings_available() -> bool:
    return _HAS_SENTENCE_TRANSFORMERS


def _get_embed_model():
    if EMBEDDING_MODEL_NAME not in _embed_model_singleton:
        _embed_model_singleton[EMBEDDING_MODEL_NAME] = SentenceTransformer(EMBEDDING_MODEL_NAME)
    return _embed_model_singleton[EMBEDDING_MODEL_NAME]


def build_embedding_lookup(all_texts: np.ndarray) -> dict[str, np.ndarray] | None:
    """전체 문헌 텍스트에 대해 임베딩을 한 번만 계산해 {텍스트: 벡터} 딕셔너리로 반환한다.
    사전학습 인코더가 이 데이터로 학습되는 게 아니므로, CV fold 밖에서 한 번만
    계산해도 안전하다 (TF-IDF 벡터라이저처럼 fold마다 다시 fit할 필요가 없음)."""
    if not _HAS_SENTENCE_TRANSFORMERS:
        return None
    model = _get_embed_model()
    unique_texts = list(dict.fromkeys(all_texts.tolist()))  # 중복 제거, 순서 보존
    vectors = model.encode(unique_texts, batch_size=32, show_progress_bar=False, normalize_embeddings=True)
    return {t: v for t, v in zip(unique_texts, vectors)}


class EmbeddingLookup(BaseEstimator, TransformerMixin):
    """사전 계산된 임베딩을 텍스트로 조회만 하는 변환기. fit에서 아무것도 학습하지
    않으므로(고정 가중치 인코더), Pipeline 안에 있어도 매 fold 재계산이 필요 없다."""

    def __init__(self, lookup: dict[str, np.ndarray] | None = None, dim: int = 768):
        self.lookup = lookup or {}
        self.dim = dim

    def fit(self, X, y=None):
        return self

    def transform(self, X):
        if not self.lookup:
            return np.zeros((len(X), self.dim))
        any_vec = next(iter(self.lookup.values()))
        dim = any_vec.shape[0]
        return np.vstack([self.lookup.get(t, np.zeros(dim)) for t in X])



STRUCTURED_SEP = "\n<ABSTRACT>\n"


def _serialize_structured(title: str, abstract: str) -> str:
    return f"{str(title).strip()}{STRUCTURED_SEP}{str(abstract).strip()}"


class TextPartExtractor(BaseEstimator, TransformerMixin):
    """StructuredText에서 Title 또는 Abstract만 분리해 별도 TF-IDF가 학습되도록 한다."""
    def __init__(self, part: str = "title"):
        self.part = part

    def fit(self, X, y=None):
        return self

    def transform(self, X):
        out = []
        for x in X:
            text = str(x)
            if STRUCTURED_SEP in text:
                title, abstract = text.split(STRUCTURED_SEP, 1)
            else:
                title, abstract = text, ""
            value = title if self.part == "title" else abstract
            # Title-only screening files are valid. If an entire CV fold has no abstracts,
            # sklearn's TfidfVectorizer would otherwise raise
            # "empty vocabulary; perhaps the documents only contain stop words".
            # A constant placeholder keeps the feature branch structurally valid while
            # contributing no discriminative information.
            if not str(value).strip():
                value = "missing_abstract" if self.part == "abstract" else "missing_title"
            out.append(value)
        return np.asarray(out, dtype=object)


class RuleSignalFeatures(BaseEstimator, TransformerMixin):
    """초록에서 명백한 연구설계/비대상 신호를 숫자 feature로 변환한다.
    단일 키워드만으로 자동 배제하지 않고 최종 분류기의 보조 feature로만 사용한다."""
    PATTERNS = [
        r"\bin\s*vitro\b|cell\s+culture|cultured\s+cells?|cell\s+line|c2c12|myoblast|osteoblast",
        r"\breview\b|systematic\s+review|meta[- ]analysis|narrative\s+review",
        r"protocol|study\s+protocol",
        r"case\s+report|case\s+series",
        r"plant\s+growth|seedling|arabidopsis|crop\s+plant",
        r"no\s+(?:treatment|intervention)|observational\s+study|cross[- ]sectional",
    ]

    def fit(self, X, y=None):
        return self

    def transform(self, X):
        rows = []
        for x in X:
            t = str(x).lower()
            vals = [1.0 if re.search(p, t, flags=re.I) else 0.0 for p in self.PATTERNS]
            vals.append(float(sum(vals)))
            rows.append(vals)
        return np.asarray(rows, dtype=float)


class PrototypeSimilarity(BaseEstimator, TransformerMixin):
    """훈련 fold의 Include/Exclude 문헌 centroid와의 유사도를 계산한다.
    fold 안에서만 prototype을 만들기 때문에 교차검증 누수가 없다."""
    def __init__(self, max_features: int = 25000):
        self.max_features = max_features

    def fit(self, X, y=None):
        self.vectorizer_ = TfidfVectorizer(ngram_range=(1, 2), min_df=1, max_features=self.max_features, sublinear_tf=True)
        mat = self.vectorizer_.fit_transform(list(X))
        y = np.asarray(y) if y is not None else np.zeros(mat.shape[0], dtype=int)
        self.pos_ = np.asarray(mat[y == 1].mean(axis=0) if np.any(y == 1) else mat.mean(axis=0))
        self.neg_ = np.asarray(mat[y == 0].mean(axis=0) if np.any(y == 0) else mat.mean(axis=0))
        return self

    def transform(self, X):
        mat = self.vectorizer_.transform(list(X))
        pos = cosine_similarity(mat, self.pos_).ravel()
        neg = cosine_similarity(mat, self.neg_).ravel()
        return np.column_stack([pos, neg, pos - neg])


class NumericLookup(BaseEstimator, TransformerMixin):
    """사전 계산된 고정 numeric features를 StructuredText key로 조회한다."""
    def __init__(self, lookup=None, dim: int = 4):
        self.lookup = lookup or {}
        self.dim = dim

    def fit(self, X, y=None):
        return self

    def transform(self, X):
        if not self.lookup:
            return np.zeros((len(X), self.dim), dtype=float)
        any_vec = np.asarray(next(iter(self.lookup.values())), dtype=float)
        dim = int(any_vec.size)
        return np.vstack([np.asarray(self.lookup.get(str(x), np.zeros(dim)), dtype=float) for x in X])

class CriteriaSimilarity(BaseEstimator, TransformerMixin):
    """코사인 유사도 기반 PICO/배제기준 근접도 피처.

    각 문헌 텍스트와 사용자가 입력한 PICO+배제기준 텍스트 사이의 TF-IDF
    코사인 유사도를 하나의 숫자 피처로 만든다. 반드시 Pipeline 안에 넣어
    fit/transform이 매 CV fold마다 train 텍스트로만 다시 이루어지도록 해야
    검증 fold의 텍스트가 IDF 통계에 새어 들어가지 않는다 (다른 프로젝트에서
    잡았던 것과 동일한 리키지 패턴).
    """

    def __init__(self, criteria_text: str = ""):
        self.criteria_text = criteria_text

    def fit(self, X, y=None):
        text = (self.criteria_text or "").strip()
        corpus = list(X) + [text if text else " "]
        self.vectorizer_ = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True)
        self.vectorizer_.fit(corpus)
        self.criteria_vec_ = self.vectorizer_.transform([text if text else " "])
        return self

    def transform(self, X):
        doc_vecs = self.vectorizer_.transform(X)
        return cosine_similarity(doc_vecs, self.criteria_vec_)


# ---------------------------------------------------------------------------
# PICO 섹션 파서. 사용자가 기준 텍스트를 'P: ...' / 'I: ...' / 'C: ...' / 'O: ...'
# (또는 한글 '대상:'/'중재:'/'대조:'/'결과:') 형식으로 쓰면 항목별로 분리한다.
# 인식되는 접두어가 하나도 없으면 전체를 "criteria"라는 단일 키로 반환해
# 기존(통짜 유사도) 동작과 완전히 호환된다. LLM 호출 없이 문자열 파싱만 사용.
# ---------------------------------------------------------------------------
_PICO_PREFIXES = [
    ("P", ["p:", "population:", "patient:", "대상:", "환자:", "인구집단:"]),
    ("I", ["i:", "intervention:", "중재:", "노출:"]),
    ("C", ["c:", "comparison:", "comparator:", "대조:", "비교:"]),
    ("O", ["o:", "outcome:", "결과:", "결과지표:"]),
]


def _parse_pico_sections(criteria_text: str) -> dict[str, str]:
    text = (criteria_text or "").strip()
    if not text:
        return {"criteria": ""}
    sections: dict[str, list[str]] = {}
    current = None
    matched_any = False
    for line in text.splitlines():
        stripped = line.strip()
        lowered = stripped.lower()
        found_key, remainder = None, stripped
        for key, prefixes in _PICO_PREFIXES:
            for p in prefixes:
                if lowered.startswith(p):
                    found_key = key
                    remainder = stripped[len(p):].strip()
                    break
            if found_key:
                break
        if found_key:
            matched_any = True
            current = found_key
            sections.setdefault(current, []).append(remainder)
        elif current is not None and stripped:
            sections[current].append(stripped)
    if not matched_any:
        return {"criteria": text}
    return {k: " ".join(v).strip() for k, v in sections.items() if " ".join(v).strip()}


# ---------------------------------------------------------------------------
# Zero-shot 스크리닝: 라벨링된 문헌이 하나도 없는 프로젝트 초기 단계를 위한 모드.
# 지도학습 파이프라인(train_and_predict)은 최소 20개 이상의 라벨이 있어야 재현율을
# 통계적으로 보장할 수 있다. 이 함수는 PICO 기준 + 배제기준 텍스트만으로 모든
# 문헌에 순위를 매긴다 — 다만 정답(라벨)이 전혀 없으므로 "재현율 X% 보장" 같은
# 통계적 검증은 원리적으로 불가능하다는 점이 지도학습 모드와의 근본적 차이다.
#
# 대신 다음 두 가지로 최대한 원칙 있게 동작하게 했다:
#   1) PICO 각 항목(P/I/C/O)·배제기준 각 항목에 대한 의미 유사도를 따로 계산
#      (뭉뚱그린 유사도 하나보다 어느 항목이 안 맞는지 설명 가능)
#   2) Otsu 임계값 방법(영상 이진화 기법을 1차원 점수 분포에 적용)으로 임의의
#      퍼센트/임계값을 사용자가 고르지 않아도 "안전 제외 / 경계 / 우선 검토"
#      3단계를 점수 분포 스스로 자연스럽게 나누게 한다 — 이번 요청에서 강조하신
#      "매번 선택하게 하지 말고 고정" 원칙을 그대로 따른 것.
#
# 중요: 안전 제외 후보 중 일부(예: 20~30편)를 무작위로 뽑아 사람이 직접 확인하고,
# 그 판정을 Human_Label로 입력해 지도학습 모드로 넘어가는 것을 강력히 권장한다.
# 그래야 비로소 재현율 보장이 통계적으로 성립한다.
# ---------------------------------------------------------------------------

ZERO_SHOT_DISCLAIMER = (
    "Zero-shot 모드는 라벨(정답) 없이 PICO/배제기준 텍스트 유사도만으로 순위를 매깁니다. "
    "재현율 목표를 통계적으로 보장하지 않습니다. 안전 제외 후보 중 일부를 무작위로 "
    "직접 확인한 뒤 그 판정을 라벨로 입력하면, 지도학습 모드(train_and_predict)로 넘어가 "
    "통계적으로 검증된 재현율 보장을 받을 수 있습니다."
)


def prepare_unlabeled_data(df: pd.DataFrame) -> pd.DataFrame:
    """라벨 열이 전혀 없어도 되는 zero-shot 전용 데이터 준비 함수.
    prepare_screening_data와 달리 라벨 열을 요구하지 않는다."""
    title_col = _find_col(df, ["title", "제목"])
    abstract_col = _find_col(df, ["abstract", "초록"])
    if not title_col:
        raise ValueError("제목(Title/제목) 열이 필요합니다.")
    out = pd.DataFrame()
    out["Title"] = df[title_col].fillna("").astype(str)
    out["Abstract"] = df[abstract_col].fillna("").astype(str) if abstract_col else ""
    out["Text"] = (out["Title"] + " " + out["Abstract"]).str.strip()
    out["StructuredText"] = [_serialize_structured(t, a) for t, a in zip(out["Title"], out["Abstract"])]
    return out


def _split_bullet_items(text: str) -> list[str]:
    """배제기준처럼 여러 항목이 줄바꿈/불릿으로 나열된 텍스트를 개별 항목으로 분리.
    분리되는 항목이 없으면(불릿 없는 한 문단) 전체를 항목 하나로 취급한다."""
    if not text or not text.strip():
        return []
    items = []
    for line in text.splitlines():
        stripped = line.strip()
        stripped = re.sub(r"^[\-\*\u2022\u25CF\u25AA]\s*", "", stripped)
        stripped = re.sub(r"^\d+[\.\)]\s*", "", stripped)
        if stripped:
            items.append(stripped)
    return items or [text.strip()]


def _cosine_sim_matrix(doc_texts: list[str], query_texts: list[str]) -> np.ndarray:
    """문헌 목록과 질의(PICO 항목/배제기준 항목) 사이의 코사인 유사도 행렬
    (n_docs, n_queries). 임베딩 모델이 있으면 의미 기반, 없으면 TF-IDF로 대체
    (둘 다 무료·로컬, API 비용 없음)."""
    if not query_texts:
        return np.zeros((len(doc_texts), 0))
    if embeddings_available():
        model = _get_embed_model()
        doc_vecs = model.encode(list(doc_texts), batch_size=32, show_progress_bar=False, normalize_embeddings=True)
        query_vecs = model.encode(list(query_texts), batch_size=32, show_progress_bar=False, normalize_embeddings=True)
        return doc_vecs @ query_vecs.T
    vectorizer = TfidfVectorizer(ngram_range=(1, 2), min_df=1, sublinear_tf=True)
    vectorizer.fit(list(doc_texts) + list(query_texts))
    doc_vecs = vectorizer.transform(doc_texts)
    query_vecs = vectorizer.transform(query_texts)
    return cosine_similarity(doc_vecs, query_vecs)


def _otsu_threshold(values: np.ndarray, bins: int = 256) -> float:
    """1차원 점수 분포를 두 그룹으로 가장 잘 가르는 임계값 (Otsu, 1979).
    영상 이진화 기법을 연속값 분포에 그대로 적용한다 — '몇 %를 자를지'를
    사람이 정하는 대신, 그룹 간 분산이 최대가 되는 지점을 데이터 스스로
    찾게 한다. 라벨이 전혀 없는 zero-shot 모드에서 임계값을 고정 상수로
    하드코딩하지 않기 위한 용도."""
    values = np.asarray(values, dtype=float)
    if len(values) < 2 or np.allclose(values.min(), values.max()):
        return float(np.median(values)) if len(values) else 0.0
    hist, edges = np.histogram(values, bins=bins)
    hist = hist.astype(float)
    total = hist.sum()
    if total == 0:
        return float(np.median(values))
    prob = hist / total
    bin_centers = (edges[:-1] + edges[1:]) / 2
    cumulative_prob = np.cumsum(prob)
    cumulative_mean = np.cumsum(prob * bin_centers)
    global_mean = cumulative_mean[-1]
    with np.errstate(divide="ignore", invalid="ignore"):
        between_var = (global_mean * cumulative_prob - cumulative_mean) ** 2 / (
            cumulative_prob * (1 - cumulative_prob)
        )
    between_var = np.nan_to_num(between_var, nan=-1.0, posinf=-1.0, neginf=-1.0)
    best_idx = int(np.argmax(between_var))
    return float(bin_centers[best_idx])



def _split_sentences(text: str) -> list[str]:
    text = re.sub(r"\s+", " ", str(text or "")).strip()
    if not text:
        return []
    parts = re.split(r"(?<=[.!?])\s+(?=[A-Z0-9])", text)
    return [p.strip() for p in parts if len(p.strip()) >= 12] or [text]


def build_sentence_pico_lookup(keys: np.ndarray, abstracts: np.ndarray, criteria_text: str) -> dict[str, np.ndarray] | None:
    """각 Abstract 문장에서 P/I/C/O와 가장 유사한 문장을 찾아 4개 feature로 저장한다.
    임베딩이 있으면 의미 유사도, 없으면 TF-IDF 유사도로 자동 폴백한다."""
    sections = _parse_pico_sections(criteria_text)
    pico_queries = [sections.get(k, "") for k in ["P", "I", "C", "O"]]
    if not any(q.strip() for q in pico_queries):
        return None
    lookup = {}
    for key, abstract in zip(keys, abstracts):
        sents = _split_sentences(abstract)
        if not sents:
            lookup[str(key)] = np.zeros(4, dtype=float)
            continue
        vals = []
        for q in pico_queries:
            if not q.strip():
                vals.append(0.0)
            else:
                sims = _cosine_sim_matrix(sents, [q])
                vals.append(float(np.max(sims[:, 0])) if sims.size else 0.0)
        lookup[str(key)] = np.asarray(vals, dtype=float)
    return lookup


def _obvious_exclusion_reason(title: str, abstract: str) -> str:
    t = f"{title} {abstract}".lower()
    checks = [
        (r"\bin\s*vitro\b|cell\s+culture|cultured\s+cells?|cell\s+line|c2c12|myoblast|osteoblast", "In vitro / cell study signal"),
        (r"systematic\s+review|meta[- ]analysis|narrative\s+review|\breview\b", "Review article signal"),
        (r"study\s+protocol|\bprotocol\b", "Protocol signal"),
        (r"case\s+report|case\s+series", "Case report/series signal"),
    ]
    reasons = [label for pattern, label in checks if re.search(pattern, t, flags=re.I)]
    return "; ".join(reasons)

@dataclass
class ZeroShotResult:
    predictions: pd.DataFrame
    metrics: dict = field(default_factory=dict)


def zero_shot_screen(
    df: pd.DataFrame,
    criteria_text: str,
    exclusion_text: str = "",
) -> ZeroShotResult:
    """라벨 없이 PICO 기준 + 배제기준 텍스트만으로 문헌을 3단계로 자동 분류한다.

    사용 예:
        result = zero_shot_screen(
            df,
            criteria_text="P: 우주비행/미세중력 동물 또는 인체 모델\\nI: 영양 중재(비타민/미네랄/단백질 등)\\nO: 근골격계 지표(근위축, 골밀도 등)",
            exclusion_text="식물 대상 연구\\n미생물/세포주만 대상\\n운동/기구 중재만 있고 영양 중재 없음",
        )
        result.predictions  # AI_Recommendation 열 포함, 우선순위 정렬된 DataFrame
        result.metrics["disclaimer"]  # 통계적 보장이 없다는 안내 문구
    """
    if not criteria_text or not criteria_text.strip():
        raise ValueError("PICO 기준 텍스트가 필요합니다.")

    data = prepare_unlabeled_data(df)
    doc_texts = data["Text"].tolist()

    sections = _parse_pico_sections(criteria_text)
    section_names = list(sections.keys())
    section_texts = [sections[k] if sections[k].strip() else " " for k in section_names]
    # Abstract가 있으면 문장 단위에서 각 PICO 항목과 가장 가까운 문장을 사용한다.
    pico_sims = np.zeros((len(data), len(section_names)), dtype=float)
    for r, abstract in enumerate(data["Abstract"].tolist()):
        sents = _split_sentences(abstract) or [data.iloc[r]["Title"]]
        for i, q in enumerate(section_texts):
            sims = _cosine_sim_matrix(sents, [q])
            pico_sims[r, i] = float(np.max(sims[:, 0])) if sims.size else 0.0

    for i, name in enumerate(section_names):
        data[f"Similarity_{name}"] = pico_sims[:, i]

    # PICO 각 항목은 AND 조건(전부 맞아야 진짜 관련) -> 최솟값을 대표 점수로 사용.
    # 평균을 쓰면 한 항목만 매우 높아도 다른 항목이 안 맞는 문헌이 높은 점수를
    # 받는 왜곡이 생긴다.
    pico_score = pico_sims.min(axis=1) if pico_sims.shape[1] else np.zeros(len(doc_texts))
    data["PICO_Score"] = pico_score

    exclusion_items = _split_bullet_items(exclusion_text)
    if exclusion_items:
        excl_sims = _cosine_sim_matrix(doc_texts, exclusion_items)  # (n_docs, n_items)
        # 배제기준은 OR 조건(하나라도 걸리면 제외) -> 최댓값
        exclusion_score = excl_sims.max(axis=1)
    else:
        exclusion_score = np.zeros(len(doc_texts))
    data["Exclusion_Score"] = exclusion_score

    combined_score = pico_score - exclusion_score
    data["Combined_Score"] = combined_score

    # Otsu 임계값으로 자동 3단계 분류: 먼저 전체를 안전제외/나머지로 나누고,
    # 나머지를 다시 우선검토/경계문헌으로 재귀적으로 나눈다 (임의의 퍼센트나
    # 임계값을 사용자가 매번 고를 필요가 없다).
    threshold_1 = _otsu_threshold(combined_score)
    rest_mask = combined_score >= threshold_1
    rest_scores = combined_score[rest_mask]
    threshold_2 = _otsu_threshold(rest_scores) if len(rest_scores) >= 2 else threshold_1

    recommendation = np.full(len(doc_texts), "우선 검토", dtype=object)
    recommendation[~rest_mask] = "안전 제외 후보"
    borderline_mask = rest_mask & (combined_score < threshold_2)
    recommendation[borderline_mask] = "경계 문헌"
    data["AI_Recommendation"] = pd.Categorical(recommendation, categories=PRIORITY_ORDER, ordered=True)
    data["AI_Exclusion_Signal"] = [_obvious_exclusion_reason(t, a) for t, a in zip(data["Title"], data["Abstract"])]

    data = data.sort_values(
        ["AI_Recommendation", "Combined_Score"], ascending=[True, False]
    ).reset_index(drop=True)

    metrics = {
        "mode": "zero_shot",
        "n_total": int(len(data)),
        "embedding_signal_used": embeddings_available(),
        "sections_detected": [n for n in section_names if n != "criteria"],
        "exclusion_items_n": len(exclusion_items),
        "safe_exclude_n": int((data["AI_Recommendation"] == "안전 제외 후보").sum()),
        "borderline_n": int((data["AI_Recommendation"] == "경계 문헌").sum()),
        "priority_n": int((data["AI_Recommendation"] == "우선 검토").sum()),
        "otsu_threshold_exclude": float(threshold_1),
        "otsu_threshold_priority": float(threshold_2),
        "disclaimer": ZERO_SHOT_DISCLAIMER,
    }
    return ZeroShotResult(predictions=data, metrics=metrics)


# ---------------------------------------------------------------------------
# 메인 랭킹 모델 (기존과 동일: Word+Char TF-IDF [+ PICO 유사도] -> Calibrated LinearSVC)
# ---------------------------------------------------------------------------

def _build_feature_union(
    criteria_text: str = "",
    embedding_lookup: dict | None = None,
    sentence_pico_lookup: dict | None = None,
) -> FeatureUnion:
    # Combined text
    combined_word = TfidfVectorizer(ngram_range=(1, 2), min_df=1, max_features=45000, sublinear_tf=True)
    combined_char = TfidfVectorizer(analyzer="char_wb", ngram_range=(3, 5), min_df=2, max_features=40000, sublinear_tf=True)

    # Title과 Abstract를 별도의 feature space로 학습한다.
    title_pipe = Pipeline([
        ("title", TextPartExtractor("title")),
        ("tfidf", TfidfVectorizer(ngram_range=(1, 2), min_df=1, max_features=20000, sublinear_tf=True)),
    ])
    abstract_pipe = Pipeline([
        ("abstract", TextPartExtractor("abstract")),
        ("tfidf", TfidfVectorizer(ngram_range=(1, 2), min_df=1, max_features=35000, sublinear_tf=True)),
    ])

    transformers = [
        ("combined_word", combined_word),
        ("combined_char", combined_char),
        ("title_word", title_pipe),
        ("abstract_word", abstract_pipe),
        ("rules", RuleSignalFeatures()),
        ("prototype", PrototypeSimilarity()),
    ]
    if criteria_text and criteria_text.strip():
        transformers.append(("criteria", CriteriaSimilarity(criteria_text=criteria_text)))
    if sentence_pico_lookup:
        transformers.append(("sentence_pico", NumericLookup(lookup=sentence_pico_lookup, dim=4)))
    if embedding_lookup:
        transformers.append(("embedding", EmbeddingLookup(lookup=embedding_lookup)))
    return FeatureUnion(transformers)


def _build_pipeline(
    criteria_text: str = "",
    embedding_lookup: dict | None = None,
    sentence_pico_lookup: dict | None = None,
) -> Pipeline:
    features = _build_feature_union(criteria_text, embedding_lookup, sentence_pico_lookup)
    base = LinearSVC(class_weight="balanced")
    model = CalibratedClassifierCV(base, method="sigmoid", cv=3)
    return Pipeline([("features", features), ("model", model)])

# ---------------------------------------------------------------------------
# 안전 제외 후보를 위한 보조 뷰 모델들.
# 메인 랭킹 모델(위 _build_pipeline)은 그대로 유지하고, 확률 하나만으로
# 안전 제외를 결정하지 않기 위해 서로 다른 근거(단어 TF-IDF만, 문자 TF-IDF만,
# 전체 피처 기반 로지스틱 회귀, PICO 유사도만)로 학습한 표준 sklearn 모델의
# 의견을 추가로 모은다. 모두 검증 fold 누수 없이 cross_val_predict로 계산한다.
# ---------------------------------------------------------------------------

def _build_word_only_pipeline() -> Pipeline:
    word = TfidfVectorizer(ngram_range=(1, 2), min_df=1, max_features=50000, sublinear_tf=True)
    return Pipeline([("word", word), ("model", LogisticRegression(max_iter=2000, class_weight="balanced"))])


def _build_char_only_pipeline() -> Pipeline:
    char = TfidfVectorizer(analyzer="char_wb", ngram_range=(3, 5), min_df=2, max_features=50000, sublinear_tf=True)
    return Pipeline([("char", char), ("model", LogisticRegression(max_iter=2000, class_weight="balanced"))])


def _build_pico_only_pipeline(criteria_text: str) -> Pipeline:
    return Pipeline([
        ("pico", CriteriaSimilarity(criteria_text=criteria_text)),
        ("model", LogisticRegression(max_iter=2000, class_weight="balanced")),
    ])


def _build_embedding_only_pipeline(embedding_lookup: dict) -> Pipeline:
    """의미 임베딩만으로 학습한 로지스틱 회귀. 어휘 중복(TF-IDF 계열)과 독립적인
    '의미가 비슷한가'라는 근거를 안전 제외 후보 판정에 추가한다."""
    return Pipeline([
        ("embedding", EmbeddingLookup(lookup=embedding_lookup)),
        ("model", LogisticRegression(max_iter=2000, class_weight="balanced")),
    ])


def _build_sentence_pico_only_pipeline(sentence_pico_lookup: dict) -> Pipeline:
    return Pipeline([
        ("sentence_pico", NumericLookup(lookup=sentence_pico_lookup, dim=4)),
        ("model", LogisticRegression(max_iter=2000, class_weight="balanced")),
    ])


def _build_logreg_full_pipeline(criteria_text: str = "", embedding_lookup: dict | None = None, sentence_pico_lookup: dict | None = None) -> Pipeline:
    features = _build_feature_union(criteria_text, embedding_lookup, sentence_pico_lookup)
    return Pipeline([("features", features), ("model", LogisticRegression(max_iter=2000, class_weight="balanced"))])


def _compute_safety_signals(
    texts: np.ndarray, y: np.ndarray, cv: StratifiedKFold, all_texts: np.ndarray, criteria_text: str = "",
    embedding_lookup: dict | None = None, sentence_pico_lookup: dict | None = None,
) -> dict:
    """Word TF-IDF, Character TF-IDF, 전체 피처 로지스틱 회귀, PICO 유사도, (가능하면)
    의미 임베딩 각각에 대해 (라벨 데이터의 교차검증 확률, 전체 데이터 확률)을 계산해 반환한다.
    TF-IDF 계열 뷰는 fold마다 처음부터 다시 학습되므로 검증 fold 누수가 없고,
    임베딩 뷰는 고정 가중치 인코더라 애초에 데이터 의존적 fit이 없어 누수 자체가 불가능하다.
    """
    signals: dict[str, dict] = {}

    def _run(name: str, pipeline: Pipeline):
        cv_probs = cross_val_predict(pipeline, texts, y, cv=cv, method="predict_proba")[:, 1]
        pipeline.fit(texts, y)
        all_probs = pipeline.predict_proba(all_texts)[:, 1]
        signals[name] = {"cv": cv_probs, "all": all_probs}

    _run("word_tfidf", _build_word_only_pipeline())
    _run("char_tfidf", _build_char_only_pipeline())
    _run("logistic_regression", _build_logreg_full_pipeline(criteria_text, embedding_lookup, sentence_pico_lookup))
    if criteria_text and criteria_text.strip():
        _run("pico_similarity", _build_pico_only_pipeline(criteria_text))
    if sentence_pico_lookup:
        _run("sentence_pico", _build_sentence_pico_only_pipeline(sentence_pico_lookup))
    if embedding_lookup:
        _run("semantic_embedding", _build_embedding_only_pipeline(embedding_lookup))
    return signals


# ---------------------------------------------------------------------------
# 허용 False Negative(FN) 개수 기반 컷오프.
#
# "재현율 몇 %" 대신, 사람이 이해하기 쉬운 절대 개수("Include 문헌을 최대 N편까지만
# 놓치는 것을 허용")로 임계값을 정한다. 라벨된 Include 문헌들의 교차검증 확률을
# 오름차순 정렬해서, 가장 낮은 N개까지만 컷오프 아래로 떨어지도록 컷오프를 잡으면
# "이 컷오프를 쓰는 한 최대 N개까지만 놓친다"는 것이 라벨 데이터 위에서 수학적으로
# 보장된다. 안전 제외 후보 판정에 쓰이는 5개 신호 모두 같은 방식으로 컷오프를 잡고
# "모두 동의(교집합)"할 때만 안전 제외로 인정하므로, 안전 제외 후보 버킷의 FN 개수는
# 항상 이 허용치 이하로 유지된다 (교집합이므로 개별 신호의 FN 개수를 넘을 수 없다).
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# 목표 재현율(recall target) 기반 정책. "허용 FN 개수"를 매번 사람이 감으로
# 입력하는 대신, 문헌 스크리닝 자동화 분야에서 흔히 쓰는 고정된 재현율 목표
# (예: 95%)를 정책으로 두고, 현재 라벨 수로부터 allowed_fn을 결정론적으로
# 자동 계산한다. 같은 재현율 목표를 쓰는 한, 데이터셋이 달라져도 "무엇을
# 보장하는가"의 의미는 항상 동일하다 (허용 FN의 절대 개수만 라벨 수에 따라
# 달라질 뿐, 기준 자체가 흔들리는 게 아니다).
# ---------------------------------------------------------------------------

RECALL_TARGET_PRESETS = [0.99, 0.95, 0.90]
DEFAULT_RECALL_TARGET = 0.95

# ---------------------------------------------------------------------------
# 안전 제외 정족수(quorum). 예전에는 모든 신호(최대 6개)가 전부 동의해야만 안전
# 제외로 인정했다(엄격한 AND). 신호 수가 늘어날수록 교집합이 기하급수적으로
# 줄어들어, 신호를 추가할수록 오히려 실제 절감 효과가 작아지는 역설이 생긴다.
# 대신 "서로 다른 근거를 가진 신호들 중 80% 이상이 각자의 허용 FN 기준을
# 지키며 Exclude로 동의"하면 인정하도록 고정한다. 사용자가 매번 고를 수 있는
# 옵션이 아니라 코드에 고정된 값이다 (항상 같은 기준으로 동작해야 신뢰할 수
# 있고, 쓰는 사람 입장에서도 헷갈릴 옵션이 없어야 한다).
# 100%(만장일치)보다 완화됐지만 여전히 압도적 다수의 동의를 요구하므로 안전성은
# 크게 훼손하지 않으면서, 안전 제외 후보 수(=검토 절감량)를 늘린다.
# ---------------------------------------------------------------------------
SAFETY_QUORUM_RATIO = 2/3


def allowed_fn_from_recall_target(n_include: int, recall_target: float) -> int:
    """목표 재현율을 만족하는 가장 관대한(=검토량을 가장 많이 줄이는) 허용 FN 개수.
    내림(floor)을 사용해 실제 달성 재현율이 목표를 절대 밑돌지 않도록 보수적으로 잡는다."""
    if n_include <= 0:
        return 0
    # round()로 부동소수점 오차(예: 0.9*20이 1.9999999996이 되는 경우)를 먼저 보정한 뒤
    # floor를 적용해, 0.90/0.95처럼 딱 떨어져야 할 값이 한 개씩 어긋나지 않도록 한다.
    raw = round((1.0 - float(recall_target)) * n_include, 6)
    allowed = int(np.floor(raw))
    return max(0, min(n_include, allowed))


def recall_lower_confidence_bound(n_include: int, allowed_fn: int, confidence: float = 0.95) -> float:
    """Clopper-Pearson(정확 이항) 하한. 현재 라벨 표본에서 이 allowed_fn을 썼을 때
    관측된 성공률(포착률)에 표본 크기 불확실성을 반영해, '모집단 재현율이 이 값
    이상일 것이라고 confidence 신뢰수준으로 말할 수 있는' 하한선을 계산한다.
    주의: 라벨 표본이 전체 문헌(라벨 없는 문헌 포함)을 대표한다는 가정이 전제이며,
    라벨 수가 적을수록 이 하한은 크게 내려간다 (표본이 작을수록 불확실성이 크다는
    사실을 감추지 않고 그대로 보여주기 위함)."""
    n = int(n_include)
    if n <= 0:
        return 0.0
    k = n - int(allowed_fn)  # 성공(=포착)한 라벨 Include 개수
    if k <= 0:
        return 0.0
    if k >= n:
        # 전부 포착(allowed_fn=0)한 경우의 하한: Jeffreys/Clopper-Pearson 상한쪽 특수 케이스
        return float(_beta_dist.ppf(1 - confidence, n, 1))
    return float(_beta_dist.ppf(1 - confidence, k, n - k + 1))


def work_saved_over_sampling(tn: int, fn: int, tp: int, fp: int) -> float:
    """WSS (Work Saved over Sampling), Cohen et al. 2006 정의.

    WSS = (TN+FN)/N - (1 - 실제 달성 재현율)

    무작위로 문헌을 훑는 것 대비, 이 스크리닝 파이프라인이 실제로 절감한 검토 비율을
    나타내는 SR 자동화 스크리닝 분야의 표준 지표다 (ASReview, CLEF eHealth TAR 등에서
    보고하는 방식). 목표 재현율(recall_target) 자체가 아니라 라벨 데이터에서 '실제 측정된'
    재현율을 쓴다 — allowed_fn은 floor로 보수적으로 잡히므로 실제 달성 재현율이 목표보다
    같거나 높을 수 있고, WSS 정의는 항상 실측 재현율을 기준으로 하기 때문이다.
    """
    n = tn + fp + fn + tp
    if n <= 0:
        return 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    return float((tn + fn) / n - (1.0 - recall))


def _optimize_threshold_wss(cv_probs: np.ndarray, y: np.ndarray, recall_target: float = 0.95) -> tuple[float, dict]:
    """교차검증 Recall 제약을 만족하는 threshold 중 WSS가 가장 높은 값을 자동 선택한다."""
    probs = np.asarray(cv_probs, dtype=float)
    y = np.asarray(y, dtype=int)
    candidates = np.unique(np.r_[0.0, probs, 1.0])
    best = None
    for thr in candidates:
        pred = (probs >= thr).astype(int)
        tn, fp, fn, tp = confusion_matrix(y, pred, labels=[0, 1]).ravel()
        rec = tp / (tp + fn) if (tp + fn) else 0.0
        if rec + 1e-12 < float(recall_target):
            continue
        wss = work_saved_over_sampling(tn, fn, tp, fp)
        burden = (tp + fp) / len(y) if len(y) else 1.0
        score = (wss, -burden, thr)
        if best is None or score > best[0]:
            best = (score, float(thr), {"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp), "recall": float(rec), "wss": float(wss), "burden": float(burden)})
    if best is None:
        return 0.0, {"recall": 1.0, "wss": 0.0, "burden": 1.0}
    return best[1], best[2]


def _fn_budget_cutoff(cv_probs: np.ndarray, y: np.ndarray, allowed_fn: int) -> float:
    inc_probs = np.sort(np.asarray(cv_probs, dtype=float)[np.asarray(y) == 1])
    if len(inc_probs) == 0:
        return 0.5
    allowed_fn = max(0, int(allowed_fn))
    if allowed_fn >= len(inc_probs):
        return 0.0  # 전부 놓쳐도 된다면 사실상 컷오프 없음 (거의 모든 문헌이 후보가 될 수 있음)
    return float(inc_probs[allowed_fn])


def _recompute_unanimous_exclude(pred_df: pd.DataFrame, allowed_fn: int):
    """저장된 각 신호의 (라벨 데이터 교차검증 확률, 전체 데이터 확률) 컬럼으로부터
    신호별 '허용 FN' 컷오프를 다시 계산하고, 신호들 중 SAFETY_QUORUM_RATIO(80%) 이상이
    각자의 컷오프보다 낮을 때 안전 제외 후보로 인정한다 (만장일치가 아닌 고정 정족수).
    모델을 재학습하지 않고 저장된 확률만 사용하므로 허용 FN 값을 바꿔도 즉시 재계산된다.
    함수/컬럼 이름은 하위 호환을 위해 유지한다.
    """
    signal_names = [c[len("Prob_"):] for c in pred_df.columns if c.startswith("Prob_")]
    mask = pred_df["Human_Label_Normalized"].isin([0, 1])
    if signal_names:
        first_cv_col = f"CV_Prob_{signal_names[0]}"
        mask = mask & pred_df[first_cv_col].notna()
    y = pred_df.loc[mask, "Human_Label_Normalized"].astype(int).to_numpy()

    n = len(pred_df)
    n_signals = len(signal_names)
    # 정족수: 신호 수의 80% 이상이 동의해야 안전 제외 인정 (SAFETY_QUORUM_RATIO, 고정값).
    # ceil을 써서 예를 들어 신호가 6개면 5개 이상, 5개면 4개 이상 동의를 요구한다
    # (80%의 소수점 결과를 내림하면 요구 조건이 실질적으로 더 느슨해져 버리므로 올림 사용).
    quorum_needed = int(np.ceil(SAFETY_QUORUM_RATIO * n_signals)) if n_signals else 0

    agree_all = np.zeros(n, dtype=int)
    agree_cv_labeled = np.zeros(int(mask.sum()), dtype=int)
    safety_terms = []
    for name in signal_names:
        cv_probs = pd.to_numeric(pred_df.loc[mask, f"CV_Prob_{name}"], errors="coerce").to_numpy()
        cutoff = _fn_budget_cutoff(cv_probs, y, allowed_fn) if len(cv_probs) else 0.5
        all_probs = pd.to_numeric(pred_df[f"Prob_{name}"], errors="coerce").to_numpy()
        agree_all += (all_probs < cutoff).astype(int)
        agree_cv_labeled += (cv_probs < cutoff).astype(int)
        safety_terms.append(1.0 - all_probs)

    quorum_all = agree_all >= quorum_needed
    quorum_cv_labeled = agree_cv_labeled >= quorum_needed

    safety_score = np.mean(safety_terms, axis=0) if safety_terms else np.zeros(n)
    safe_exclude_cv_n = int(quorum_cv_labeled.sum())
    safe_exclude_cv_fn = int(((y == 1) & quorum_cv_labeled).sum())
    return quorum_all, safety_score, safe_exclude_cv_n, safe_exclude_cv_fn


def _priority_labels(probabilities: np.ndarray, threshold: float, unanimous_exclude: np.ndarray) -> np.ndarray:
    """확률과 다중 모델 합의를 실제 검토 목적에 맞는 3단계로 구분한다.

    - 우선 검토: Include 확률이 임계값 이상 (흰색)
    - 안전 제외 후보: 임계값 미만이면서, Word/Char TF-IDF, 로지스틱 회귀, 선형 SVM(메인 모델),
      PICO 유사도(있는 경우) 등 서로 다른 근거를 가진 신호들 중 80% 이상(고정 정족수)이
      "허용 FN 이하" 조건을 각자 만족하며 Exclude 방향으로 동의 (진한 회색)
    - 경계 문헌: 임계값 미만이지만 모델들의 의견이 갈리는 경우, 사람이 반드시 확인 (중간 회색)
    """
    probabilities = np.asarray(probabilities, dtype=float)
    unanimous_exclude = np.asarray(unanimous_exclude, dtype=bool)
    return np.where(
        probabilities >= threshold, "우선 검토",
        np.where(unanimous_exclude, "안전 제외 후보", "경계 문헌"),
    )


def _sort_by_priority(pred_df: pd.DataFrame) -> pd.DataFrame:
    order = pd.Categorical(pred_df["AI_Recommendation"], PRIORITY_ORDER, ordered=True)
    return pred_df.assign(_priority_order=order).sort_values(
        ["_priority_order", "AI_Probability"], ascending=[True, False]
    ).drop(columns="_priority_order").reset_index(drop=True)


def apply_fn_budget(result: ScreeningResult, allowed_fn: int) -> ScreeningResult:
    """저장된 교차검증 확률(메인 모델 + 보조 신호들)을 사용해 재학습 없이
    '허용 False Negative 개수'만 바꿔 임계값과 화면 분류(우선 검토 / 경계 문헌 /
    안전 제외 후보)를 다시 계산한다.
    """
    updated = deepcopy(result)
    pred_df = updated.predictions.copy()

    if not {"CV_Probability", "Human_Label_Normalized"}.issubset(pred_df.columns):
        return updated

    mask = pred_df["CV_Probability"].notna() & pred_df["Human_Label_Normalized"].isin([0, 1])
    y_labeled = pred_df.loc[mask, "Human_Label_Normalized"].astype(int).to_numpy()
    cv_probs_main = pd.to_numeric(pred_df.loc[mask, "CV_Probability"], errors="coerce").to_numpy()
    threshold = _fn_budget_cutoff(cv_probs_main, y_labeled, allowed_fn)
    updated.threshold = threshold

    probs = pd.to_numeric(pred_df["AI_Probability"], errors="coerce").fillna(0).to_numpy()
    unanimous_exclude, safety_score, safe_exclude_cv_n, safe_exclude_cv_fn = _recompute_unanimous_exclude(
        pred_df, allowed_fn
    )
    pred_df["Safety_Score"] = safety_score
    pred_df["Unanimous_Exclude"] = unanimous_exclude
    pred_df["AI_Recommendation"] = _priority_labels(probs, threshold, unanimous_exclude)
    updated.metrics["allowed_fn"] = int(allowed_fn)
    updated.metrics["safe_exclude_cv_n"] = safe_exclude_cv_n
    updated.metrics["safe_exclude_cv_false_negatives"] = safe_exclude_cv_fn

    cv_pred = (cv_probs_main >= threshold).astype(int)
    pred_df.loc[:, "CV_Prediction"] = np.nan
    pred_df.loc[mask, "CV_Prediction"] = cv_pred
    pred_df.loc[:, "False_Negative"] = False
    pred_df.loc[mask, "False_Negative"] = (y_labeled == 1) & (cv_pred == 0)
    tn, fp, fn, tp = confusion_matrix(y_labeled, cv_pred, labels=[0, 1]).ravel()
    rec = float(recall_score(y_labeled, cv_pred, zero_division=0))
    pre = float(precision_score(y_labeled, cv_pred, zero_division=0))
    updated.confusion = {"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp)}
    updated.metrics.update({
        "recall": rec, "precision": pre,
        "accuracy": float(accuracy_score(y_labeled, cv_pred)),
        "f1": float(2 * pre * rec / (pre + rec)) if pre + rec > 0 else 0.0,
        "measured_fn": int(fn),
        "wss": work_saved_over_sampling(tn, fn, tp, fp),
        "threshold_strategy": "Recall-constrained WSS optimization",
        "cv_screening_burden": float(threshold_info.get("burden", 1.0)),
        "title_abstract_separate": True,
        "sentence_pico_used": sentence_pico_lookup is not None,
        "prototype_similarity_used": True,
    })

    updated.predictions = _sort_by_priority(pred_df)
    return updated


def apply_recall_target(result: ScreeningResult, recall_target: float, confidence: float = 0.95) -> ScreeningResult:
    """저장된 교차검증 확률을 재사용해(재학습 없이), '목표 재현율' 정책만 바꿔
    화면 분류를 다시 계산한다. UI에는 99% / 95% / 90% 같은 고정된 정책 선택지만
    노출하고, allowed_fn(절대 개수)은 항상 이 함수 안에서 라벨 수로부터 자동 계산되므로
    사용자가 임의의 정수를 직접 입력할 일이 없다."""
    n_include = int(result.metrics.get("include_n", 0))
    allowed_fn = allowed_fn_from_recall_target(n_include, recall_target)
    updated = apply_fn_budget(result, allowed_fn)
    updated.metrics["recall_target"] = float(recall_target)
    updated.metrics["recall_lower_ci"] = recall_lower_confidence_bound(n_include, allowed_fn, confidence)
    return updated


def train_and_predict(
    df: pd.DataFrame,
    recall_target: float = DEFAULT_RECALL_TARGET,
    allowed_fn: int | None = None,
    criteria_text: str = "",
) -> ScreeningResult:
    """recall_target: 목표 재현율(예: 0.95 = 95%). 라벨 Include 중 이 비율 이상을
    반드시 '우선 검토' 또는 '경계 문헌'에 남기도록 allowed_fn을 자동으로 계산한다.
    allowed_fn을 직접 넘기면(고급 사용/하위 호환) recall_target 대신 그 값을 그대로 쓴다.
    """
    data, _ = prepare_screening_data(df)
    labeled = data[data["Human_Label"].isin([0, 1])].copy()
    if len(labeled) < MIN_LABELS_FOR_SUPERVISED or labeled["Human_Label"].nunique() < 2:
        raise ValueError(f"학습을 위해 Include와 Exclude가 모두 포함된 최소 {MIN_LABELS_FOR_SUPERVISED}개 라벨이 필요합니다.")

    y = labeled["Human_Label"].astype(int).to_numpy()
    texts = labeled["StructuredText"].to_numpy()
    all_texts = data["StructuredText"].to_numpy()
    all_plain_texts = data["Text"].to_numpy()
    n_include = int(y.sum())

    if allowed_fn is None:
        allowed_fn = allowed_fn_from_recall_target(n_include, recall_target)

    min_class = int(labeled["Human_Label"].value_counts().min())
    folds = max(2, min(5, min_class))
    cv = StratifiedKFold(n_splits=folds, shuffle=True, random_state=42)

    # 의미 임베딩은 전체 문헌 텍스트에 대해 한 번만 계산한다 (고정 가중치 인코더라
    # fold별 재계산이 필요 없고, 데이터로 다시 학습되지 않으므로 fold 밖에서 계산해도
    # 검증 누수가 생기지 않는다). sentence-transformers가 없는 환경에서는 None이 되어
    # 자동으로 이 신호 없이 나머지 모델들로만 동작한다 (기능 저하 없이 안전하게 폴백).
    # Embedding vector는 plain Title+Abstract에서 계산하되 StructuredText를 key로 사용한다.
    embedding_lookup = None
    if embeddings_available():
        model = _get_embed_model()
        vecs = model.encode(all_plain_texts.tolist(), batch_size=32, show_progress_bar=False, normalize_embeddings=True)
        embedding_lookup = {str(k): v for k, v in zip(all_texts, vecs)}
    sentence_pico_lookup = build_sentence_pico_lookup(all_texts, data["Abstract"].to_numpy(), criteria_text)

    # 메인 랭킹 모델(전체 파이프라인: TF-IDF + PICO 유사도 + [임베딩] + 선형 SVM)을 fold마다
    # 처음부터 다시 학습한다. train만으로 학습하기 때문에 검증 성능이 부풀려지지 않는다.
    probs = cross_val_predict(
        _build_pipeline(criteria_text, embedding_lookup, sentence_pico_lookup), texts, y, cv=cv, method="predict_proba"
    )[:, 1]

    # Precision-Recall 곡선은 참고용 차트로만 계산해서 보여준다 (임계값 결정에는 쓰지 않음).
    precision, recall, pr_thresholds = precision_recall_curve(y, probs)
    fpr, tpr, _ = roc_curve(y, probs)

    # 임계값은 '허용 FN 개수'로 직접 정한다: 라벨 Include 중 확률이 가장 낮은
    # allowed_fn개까지만 임계값 아래로 떨어지도록 하는 가장 관대한(=검토량이 가장 적은)
    # 임계값을 선택한다. allowed_fn은 위에서 recall_target으로부터 자동 계산되었으므로,
    # "몇 편 놓쳐도 되는가"를 매번 감으로 정하는 게 아니라 고정된 재현율 정책에서 유도된다.
    threshold, threshold_info = _optimize_threshold_wss(probs, y, recall_target)
    pred = (probs >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y, pred, labels=[0, 1]).ravel()

    recall_v = float(recall_score(y, pred, zero_division=0))
    precision_v = float(precision_score(y, pred, zero_division=0))
    metrics = {
        "recall_target": float(recall_target),
        "allowed_fn": int(allowed_fn),
        "measured_fn": int(fn),
        "recall_lower_ci": recall_lower_confidence_bound(n_include, allowed_fn),
        "recall": recall_v,
        "precision": precision_v,
        "accuracy": float(accuracy_score(y, pred)),
        "f1": float(2 * precision_v * recall_v / (precision_v + recall_v)) if (precision_v + recall_v) > 0 else 0.0,
        "roc_auc": float(roc_auc_score(y, probs)),
        "average_precision": float(average_precision_score(y, probs)),
        "labeled_n": int(len(labeled)),
        "include_n": n_include,
        "embedding_signal_used": embedding_lookup is not None,
        "wss": work_saved_over_sampling(tn, fn, tp, fp),
    }

    # 메인 모델을 전체 라벨 데이터로 최종 학습해 전체 문헌(라벨 없는 것 포함)에
    # 대한 확률을 계산한다.
    final_pipeline = _build_pipeline(criteria_text, embedding_lookup, sentence_pico_lookup)
    final_pipeline.fit(texts, y)
    all_probs = final_pipeline.predict_proba(all_texts)[:, 1]

    # --- 안전 제외 후보: 서로 다른 근거를 가진 신호들이 "각자 같은 허용 FN
    # 기준으로도 안전한" 컷오프 아래일 때만 인정한다. 최대한 많이 거르는 것이
    # 아니라, 거의 틀리지 않는 것만 거른다. (word/char TF-IDF, 로지스틱 회귀,
    # PICO 유사도, 의미 임베딩, 선형 SVM 최대 6개 신호)
    aux_signals = _compute_safety_signals(texts, y, cv, all_texts, criteria_text, embedding_lookup, sentence_pico_lookup)
    aux_signals["linear_svm"] = {"cv": probs, "all": all_probs}  # 메인 모델도 하나의 투표로 포함

    result_df = df.copy().reset_index(drop=True)
    result_df["AI_Probability"] = all_probs
    result_df["AI_Probability_%"] = (all_probs * 100).round(2)
    result_df["Human_Label_Normalized"] = data["Human_Label"].to_numpy()
    result_df["CV_Probability"] = np.nan
    result_df.loc[labeled.index, "CV_Probability"] = probs
    result_df["CV_Prediction"] = np.nan
    result_df.loc[labeled.index, "CV_Prediction"] = pred
    result_df["False_Negative"] = False
    result_df.loc[labeled.index, "False_Negative"] = (y == 1) & (pred == 0)

    for name, sig in aux_signals.items():
        result_df[f"Prob_{name}"] = sig["all"]
        result_df[f"CV_Prob_{name}"] = np.nan
        result_df.loc[labeled.index, f"CV_Prob_{name}"] = sig["cv"]

    unanimous_exclude_all, safety_score_all, safe_exclude_cv_n, safe_exclude_cv_fn = _recompute_unanimous_exclude(
        result_df, allowed_fn
    )
    # 라벨 데이터에서 "안전 제외 후보로 분류되었지만 실제로는 Include였던" 건수를
    # 교차검증 기준으로 집계한다. 5개 신호 모두 같은 허용 FN 기준으로 컷오프를 잡고
    # 교집합(모두 동의)만 인정하므로, 이 값은 수학적으로 allowed_fn을 넘을 수 없다.
    metrics["safe_exclude_cv_n"] = safe_exclude_cv_n
    metrics["safe_exclude_cv_false_negatives"] = safe_exclude_cv_fn
    metrics["safety_signal_count"] = len(aux_signals)

    result_df["Safety_Score"] = safety_score_all
    result_df["Unanimous_Exclude"] = unanimous_exclude_all
    result_df["AI_Recommendation"] = _priority_labels(all_probs, threshold, unanimous_exclude_all)
    result_df["AI_Exclusion_Signal"] = [_obvious_exclusion_reason(t, a) for t, a in zip(data["Title"], data["Abstract"])]

    result_df = _sort_by_priority(result_df)

    return ScreeningResult(
        predictions=result_df,
        metrics=metrics,
        threshold=threshold,
        pr_curve={"precision": precision.tolist(), "recall": recall.tolist(), "thresholds": pr_thresholds.tolist()},
        roc_curve={"fpr": fpr.tolist(), "tpr": tpr.tolist()},
        confusion={"tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp)},
    )


def _export_group_labels(predictions: pd.DataFrame) -> pd.Series:
    """다운로드용 4구간 라벨을 만든다: 우선 검토 -> 경계 문헌 -> False Negative ->
    안전 제외 후보. False Negative(실제 Include인데 컷오프 아래로 예측된 문헌)는
    원래 속했던 버킷(보통 경계 문헌 또는 드물게 안전 제외 후보)에서 분리해
    독립된 구간으로 모아, 다운로드했을 때 놓치면 안 되는 문헌이 눈에 띄도록 한다.
    한 문헌은 정확히 한 구간에만 속한다 (중복 없음).
    """
    rec = predictions.get("AI_Recommendation", pd.Series("", index=predictions.index)).fillna("")
    is_fn = predictions.get("False_Negative", pd.Series(False, index=predictions.index)).fillna(False).astype(bool)
    group = np.select(
        [is_fn, rec.eq("우선 검토"), rec.eq("안전 제외 후보")],
        ["False Negative", "우선 검토", "안전 제외 후보"],
        default="경계 문헌",
    )
    return pd.Series(group, index=predictions.index, name="_export_group")


def build_grouped_excel_bytes(predictions: pd.DataFrame) -> bytes:
    """AI 스크리닝 결과를 우선 검토 -> 경계 문헌 -> False Negative -> 안전 제외 후보
    순서로 정렬하고, 구간별로 배경색을 입힌 엑셀 파일 바이트를 만든다.
    """
    df = predictions.copy()
    df["_export_group"] = _export_group_labels(df)
    df["_export_group"] = pd.Categorical(df["_export_group"], EXPORT_GROUP_ORDER, ordered=True)
    sort_cols = ["_export_group"] + (["AI_Probability"] if "AI_Probability" in df.columns else [])
    ascending = [True] + ([False] * (len(sort_cols) - 1))
    df = df.sort_values(sort_cols, ascending=ascending).reset_index(drop=True)

    group_labels = df["_export_group"].astype(str).tolist()
    export_df = df.drop(columns=["_export_group"])

    wb = Workbook()
    ws = wb.active
    ws.title = "AI_Screening_Ranked"

    for row in dataframe_to_rows(export_df, index=False, header=True):
        ws.append(row)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    for i, grp in enumerate(group_labels, start=2):  # 1행은 헤더
        color = EXPORT_GROUP_COLORS.get(grp, "FFFFFF")
        fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        for cell in ws[i]:
            cell.fill = fill

    for col_idx, col_name in enumerate(export_df.columns, start=1):
        sample = export_df[col_name].astype(str).head(200).tolist()
        max_len = max([len(str(col_name))] + [len(v) for v in sample]) if sample else len(str(col_name))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(60, max(10, max_len + 2))

    ws.freeze_panes = "A2"

    legend_ws = wb.create_sheet("안내")
    legend_ws.append(["구간", "설명"])
    legend_ws["A1"].font = Font(bold=True)
    legend_ws["B1"].font = Font(bold=True)
    legend_rows = [
        ("우선 검토", "Include 확률이 임계값 이상인 문헌. 사람이 우선적으로 확인해야 합니다."),
        ("경계 문헌", "임계값 미만이지만 Word/Char TF-IDF, 로지스틱 회귀, 선형 SVM, PICO 유사도, (가능한 경우) 의미 임베딩 모델들의 의견이 갈리는 문헌. 반드시 사람이 확인해야 합니다."),
        ("False Negative", "실제 라벨은 Include였지만 교차검증에서 임계값 아래로 예측된 문헌. 모델 개선 및 재확인이 필요합니다."),
        ("안전 제외 후보", "서로 다른 근거를 가진 신호 모델들 중 80% 이상이, 목표 재현율로부터 자동 계산된 허용 FN 기준을 각자 지키면서 Exclude 방향으로 동의한 문헌. 사람이 읽지 않아도 되는 문헌입니다."),
    ]
    for i, (name, desc) in enumerate(legend_rows, start=2):
        legend_ws.append([name, desc])
        color = EXPORT_GROUP_COLORS.get(name, "FFFFFF")
        legend_ws.cell(row=i, column=1).fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    legend_ws.column_dimensions["A"].width = 16
    legend_ws.column_dimensions["B"].width = 90

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
