#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""Reusable high-recall title/abstract triage engine for systematic reviews.

The engine intentionally treats model-negative records as lower-priority
screening candidates, not as final exclusions. Human decisions always win.
"""

from __future__ import annotations

import hashlib
import io
import json
import math
import os
import re
import unicodedata
import warnings
import zipfile
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Iterable, List, Optional, Sequence, Tuple

os.environ.setdefault("MPLCONFIGDIR", str(Path(__file__).resolve().parent / ".mplcache"))
warnings.filterwarnings(
    "ignore", category=FutureWarning, message="Starting with pandas version 3.0"
)

import joblib
import matplotlib
matplotlib.use("Agg")  # headless/thread-safe backend; app.py runs this in a Streamlit worker thread
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy.sparse import csr_matrix, hstack
from sklearn.cluster import MiniBatchKMeans
from sklearn.decomposition import TruncatedSVD
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import (
    average_precision_score,
    brier_score_loss,
    confusion_matrix,
    precision_recall_curve,
    precision_recall_fscore_support,
    recall_score,
    roc_auc_score,
    roc_curve,
)
from sklearn.model_selection import StratifiedGroupKFold, StratifiedKFold


TITLE_CANDIDATES = [
    "title", "article title", "document title", "논문제목", "제목", "ti",
]
ABSTRACT_CANDIDATES = [
    "abstract", "abstract text", "초록", "요약", "ab",
]
YEAR_CANDIDATES = [
    "year", "publication year", "published year", "연도", "py",
]
LABEL_CANDIDATES = [
    "human_decision", "reviewer_decision", "final_decision", "decision",
    "label", "include/exclude", "include_exclude", "성민", "판정", "결정",
]

INCLUDE_MARKS = {
    "1", "o", "○", "include", "included", "in", "yes", "y", "포함",
}
EXCLUDE_MARKS = {
    "0", "x", "×", "exclude", "excluded", "ex", "no", "n", "배제",
}

ZONE_REVIEW = "INCLUDE_REVIEW_FIRST"
ZONE_MANUAL = "MANUAL_REVIEW"
ZONE_LOW = "LOW_PRIORITY_EXCLUDE_AUDIT"
ZONE_HUMAN_INCLUDE = "HUMAN_INCLUDE"
ZONE_HUMAN_EXCLUDE = "HUMAN_EXCLUDE"


@dataclass
class RunConfig:
    project_name: str
    population: str = ""
    intervention: str = ""
    comparator: str = ""
    outcomes: str = ""
    inclusion_criteria: str = ""
    exclusion_criteria: str = ""
    target_recall: float = 0.98
    safe_recall: float = 0.995
    cv_folds: int = 5
    audit_size: int = 50
    random_seed: int = 20260714
    min_per_class: int = 10

    def criteria_blocks(self) -> List[Tuple[str, str]]:
        blocks: List[Tuple[str, str]] = []
        for name, value in (
            ("P_Population", self.population),
            ("I_Intervention", self.intervention),
            ("C_Comparator", self.comparator),
            ("O_Outcomes", self.outcomes),
            ("Inclusion", self.inclusion_criteria),
        ):
            if clean_text(value):
                blocks.append((name, clean_text(value)))
        return blocks

    def exclusion_lines(self) -> List[str]:
        lines = re.split(r"[\r\n]+", self.exclusion_criteria or "")
        return [clean_text(x.lstrip("-•0123456789.) ")) for x in lines if clean_text(x)]


@dataclass
class PipelineResult:
    status: str
    run_dir: Path
    excel_path: Path
    zip_path: Path
    dashboard_path: Optional[Path]
    threshold_path: Optional[Path]
    model_path: Optional[Path]
    metrics: Dict[str, object]
    counts: Dict[str, int]
    message: str


def clean_text(value: object) -> str:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return ""
    text = unicodedata.normalize("NFKC", str(value)).replace("\u00a0", " ")
    return re.sub(r"\s+", " ", text).strip()


def normalize_title(value: object) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9가-힣]+", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def parse_label(value: object) -> float:
    if value is None or (isinstance(value, float) and np.isnan(value)):
        return np.nan
    if isinstance(value, (int, np.integer)) and value in (0, 1):
        return float(value)
    if isinstance(value, (float, np.floating)) and value in (0.0, 1.0):
        return float(value)
    mark = clean_text(value).lower()
    if mark in INCLUDE_MARKS:
        return 1.0
    if mark in EXCLUDE_MARKS:
        return 0.0
    return np.nan


def _column_key(value: object) -> str:
    return re.sub(r"[^a-z0-9가-힣]+", "", clean_text(value).lower())


def detect_column(columns: Iterable[object], candidates: Sequence[str]) -> Optional[str]:
    original = [str(x) for x in columns]
    keyed = {_column_key(x): x for x in original}
    for candidate in candidates:
        hit = keyed.get(_column_key(candidate))
        if hit is not None:
            return hit
    for candidate in candidates:
        ck = _column_key(candidate)
        for col in original:
            if ck and ck in _column_key(col):
                return col
    return None


def _decode_bytes(raw: bytes) -> str:
    last_error: Optional[Exception] = None
    for encoding in ("utf-8-sig", "utf-8", "cp949", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"텍스트 인코딩을 확인할 수 없습니다: {last_error}")


def parse_ris(raw: bytes) -> pd.DataFrame:
    text = _decode_bytes(raw)
    records: List[Dict[str, object]] = []
    current: Dict[str, object] = {}
    continuation_key: Optional[str] = None
    tag_map = {
        "TI": "Title", "T1": "Title", "CT": "Title",
        "AB": "Abstract", "N2": "Abstract",
        "PY": "Year", "Y1": "Year", "DA": "Year",
        "DO": "DOI", "UR": "URL", "AN": "Accession",
    }
    for line in text.splitlines():
        match = re.match(r"^([A-Z0-9]{2})\s*-\s?(.*)$", line)
        if match:
            tag, value = match.groups()
            if tag == "ER":
                if current:
                    records.append(current)
                current, continuation_key = {}, None
                continue
            key = tag_map.get(tag)
            if key:
                if key in current and value:
                    current[key] = f"{current[key]} {value}".strip()
                else:
                    current[key] = value
                continuation_key = key
            else:
                continuation_key = None
        elif continuation_key and line.strip():
            current[continuation_key] = f"{current.get(continuation_key, '')} {line.strip()}".strip()
    if current:
        records.append(current)
    if not records:
        raise ValueError("RIS 레코드를 찾지 못했습니다.")
    return pd.DataFrame(records)


def read_file_bytes(raw: bytes, filename: str) -> Dict[str, pd.DataFrame]:
    suffix = Path(filename).suffix.lower()
    bio = io.BytesIO(raw)
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        book = pd.ExcelFile(bio)
        return {name: pd.read_excel(book, sheet_name=name) for name in book.sheet_names}
    if suffix == ".csv":
        text = _decode_bytes(raw)
        try:
            frame = pd.read_csv(io.StringIO(text), sep=None, engine="python")
        except Exception:
            frame = pd.read_csv(io.StringIO(text))
        return {"Data": frame}
    if suffix == ".tsv":
        return {"Data": pd.read_csv(io.StringIO(_decode_bytes(raw)), sep="\t")}
    if suffix == ".ris":
        return {"RIS": parse_ris(raw)}
    if suffix == ".txt":
        text = _decode_bytes(raw)
        lines = [x.strip() for x in text.splitlines() if x.strip()]
        if lines and ("\t" in lines[0] or "," in lines[0]):
            try:
                return {"Data": pd.read_csv(io.StringIO(text), sep=None, engine="python")}
            except Exception:
                pass
        return {"Titles": pd.DataFrame({"Title": lines})}
    raise ValueError("지원 형식: .xlsx, .xlsm, .xls, .csv, .tsv, .ris, .txt")


def prepare_records(
    frame: pd.DataFrame,
    title_col: str,
    abstract_col: Optional[str] = None,
    year_col: Optional[str] = None,
    label_col: Optional[str] = None,
) -> pd.DataFrame:
    if title_col not in frame.columns:
        raise ValueError("선택한 제목 열을 찾을 수 없습니다.")
    out = frame.copy().reset_index(drop=True)
    out.insert(0, "_source_row", np.arange(2, len(out) + 2))
    out["Title"] = out[title_col].map(clean_text)
    out["Abstract"] = out[abstract_col].map(clean_text) if abstract_col else ""
    out["Year"] = out[year_col] if year_col else ""
    out["Human_Label"] = out[label_col].map(parse_label) if label_col else np.nan
    out["_normalized_title"] = out["Title"].map(normalize_title)
    out["_text"] = (out["Title"] + " [SEP] " + out["Abstract"]).str.strip()
    out["_group"] = out["_normalized_title"].map(
        lambda x: hashlib.sha1(x.encode("utf-8")).hexdigest()[:16] if x else ""
    )
    return out


def merge_external_labels(
    records: pd.DataFrame,
    labels: Optional[pd.DataFrame],
    title_col: Optional[str],
    label_col: Optional[str],
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    out = records.copy()
    audit_rows: List[Dict[str, object]] = []
    if labels is None or not title_col or not label_col or title_col == label_col:
        return out, pd.DataFrame(audit_rows)
    lab = labels[[title_col, label_col]].copy()
    lab["_normalized_title"] = lab[title_col].map(normalize_title)
    lab["_external_label"] = lab[label_col].map(parse_label)
    lab = lab[(lab["_normalized_title"] != "") & lab["_external_label"].notna()]
    conflict = lab.groupby("_normalized_title")["_external_label"].nunique()
    conflict_titles = set(conflict[conflict > 1].index)
    if conflict_titles:
        for title in sorted(conflict_titles):
            audit_rows.append({
                "Issue": "Conflicting external labels",
                "Normalized_Title": title,
                "Action": "Not imported",
            })
    lab = lab[~lab["_normalized_title"].isin(conflict_titles)]
    mapping = lab.drop_duplicates("_normalized_title", keep="last").set_index(
        "_normalized_title"
    )["_external_label"]
    ext = out["_normalized_title"].map(mapping)
    both = out["Human_Label"].notna() & ext.notna()
    disagreements = both & (out["Human_Label"] != ext)
    for idx in out.index[disagreements]:
        audit_rows.append({
            "Issue": "Main/external label conflict",
            "Normalized_Title": out.at[idx, "_normalized_title"],
            "Action": "Main file label retained",
        })
    fill = out["Human_Label"].isna() & ext.notna()
    out.loc[fill, "Human_Label"] = ext[fill]
    audit_rows.append({
        "Issue": "External labels imported",
        "Normalized_Title": "",
        "Action": int(fill.sum()),
    })
    return out, pd.DataFrame(audit_rows)


def validate_records(records: pd.DataFrame) -> pd.DataFrame:
    issues = []
    empty = int((records["Title"] == "").sum())
    duplicates = int(records.loc[records["_normalized_title"] != "", "_normalized_title"].duplicated().sum())
    invalid = int(records["Human_Label"].isna().sum())
    includes = int((records["Human_Label"] == 1).sum())
    excludes = int((records["Human_Label"] == 0).sum())
    for item, value, note in (
        ("Total rows", len(records), ""),
        ("Empty titles", empty, "학습·예측에서 제외"),
        ("Duplicate normalized titles", duplicates, "교차검증 시 같은 fold로 묶음"),
        ("Human include labels", includes, ""),
        ("Human exclude labels", excludes, ""),
        ("Unlabeled rows", invalid, "AI 예측 대상"),
    ):
        issues.append({"Check": item, "Count": value, "Note": note})
    return pd.DataFrame(issues)


def _word_vectorizer(n_rows: int) -> TfidfVectorizer:
    return TfidfVectorizer(
        lowercase=True,
        strip_accents="unicode",
        ngram_range=(1, 2),
        min_df=1 if n_rows < 100 else 2,
        max_df=0.995,
        max_features=60000,
        sublinear_tf=True,
        token_pattern=r"(?u)\b[\w\-]{2,}\b",
    )


def _char_vectorizer(n_rows: int) -> TfidfVectorizer:
    return TfidfVectorizer(
        lowercase=True,
        analyzer="char_wb",
        ngram_range=(3, 5),
        min_df=1 if n_rows < 100 else 2,
        max_features=70000,
        sublinear_tf=True,
    )


def _classifier(seed: int) -> LogisticRegression:
    return LogisticRegression(
        C=2.0,
        class_weight="balanced",
        solver="liblinear",
        max_iter=3000,
        random_state=seed,
    )


def _folds(y: np.ndarray, groups: np.ndarray, requested: int, seed: int):
    counts = np.bincount(y.astype(int), minlength=2)
    unique_groups = len(np.unique(groups))
    has_duplicates = unique_groups < len(groups)
    if has_duplicates:
        group_frame = pd.DataFrame({"group": groups, "label": y}).drop_duplicates("group")
        group_counts = group_frame["label"].value_counts()
        min_group_class = int(min(group_counts.get(0, 0), group_counts.get(1, 0)))
        n_splits = max(2, min(requested, min_group_class))
    else:
        n_splits = max(2, min(requested, int(counts.min())))
    if unique_groups >= n_splits and has_duplicates:
        splitter = StratifiedGroupKFold(n_splits=n_splits, shuffle=True, random_state=seed)
        return list(splitter.split(np.zeros(len(y)), y, groups))
    splitter = StratifiedKFold(n_splits=n_splits, shuffle=True, random_state=seed)
    return list(splitter.split(np.zeros(len(y)), y))


def _criteria_matrix(
    train_texts: Sequence[str],
    other_texts: Sequence[str],
    config: RunConfig,
) -> Tuple[np.ndarray, np.ndarray, Dict[str, object]]:
    blocks = config.criteria_blocks()
    exclusions = [(f"Exclusion_{i+1}", x) for i, x in enumerate(config.exclusion_lines())]
    criteria = blocks + exclusions
    if not criteria:
        return (
            np.zeros((len(train_texts), 1), dtype=float),
            np.zeros((len(other_texts), 1), dtype=float),
            {"vectorizer": None, "criteria": [], "feature_names": ["No_criteria"]},
        )
    all_text = list(train_texts) + list(other_texts)
    vec = TfidfVectorizer(
        lowercase=True,
        strip_accents="unicode",
        ngram_range=(1, 2),
        min_df=1,
        sublinear_tf=True,
        token_pattern=r"(?u)\b[\w\-]{2,}\b",
    )
    matrix = vec.fit_transform(all_text + [x[1] for x in criteria])
    docs = matrix[: len(all_text)]
    crit = matrix[len(all_text) :]
    sims = (docs @ crit.T).toarray()
    train_features = sims[: len(train_texts)]
    other_features = sims[len(train_texts) :]
    feature_names = [x[0] for x in criteria]
    return train_features, other_features, {
        "vectorizer": vec,
        "criteria": criteria,
        "feature_names": feature_names,
    }


def _metric_dict(y: np.ndarray, p: np.ndarray, threshold: float, fold_ids: np.ndarray) -> Dict[str, float]:
    pred = (p >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y, pred, labels=[0, 1]).ravel()
    precision, recall, f1, _ = precision_recall_fscore_support(
        y, pred, average="binary", zero_division=0
    )
    specificity = tn / (tn + fp) if (tn + fp) else np.nan
    npv = tn / (tn + fn) if (tn + fn) else np.nan
    fold_recalls = []
    for fold in sorted(np.unique(fold_ids)):
        mask = fold_ids == fold
        if np.any(y[mask] == 1):
            fold_recalls.append(recall_score(y[mask], pred[mask], zero_division=0))
    return {
        "threshold": float(threshold),
        "recall": float(recall),
        "min_fold_recall": float(min(fold_recalls) if fold_recalls else recall),
        "precision": float(precision),
        "f1": float(f1),
        "specificity": float(specificity),
        "npv": float(npv),
        "false_negative_rate": float(1 - recall),
        "screening_burden": float(pred.mean()),
        "work_saved_fraction": float(1 - pred.mean()),
        "tn": int(tn), "fp": int(fp), "fn": int(fn), "tp": int(tp),
    }


def choose_threshold(
    y: np.ndarray,
    p: np.ndarray,
    fold_ids: np.ndarray,
    target_recall: float,
    fold_tolerance: float = 0.05,
    max_operating_threshold: Optional[float] = 0.50,
) -> Tuple[float, pd.DataFrame]:
    candidates = np.unique(np.concatenate([
        np.array([0.0, 1.0]),
        np.linspace(0.0, 1.0, 401),
        p,
    ]))
    rows = [_metric_dict(y, p, float(t), fold_ids) for t in candidates]
    table = pd.DataFrame(rows)
    min_fold_target = max(0.0, target_recall - fold_tolerance)
    feasible = table[
        (table["recall"] >= target_recall - 1e-12)
        & (table["min_fold_recall"] >= min_fold_target - 1e-12)
    ]
    # A high threshold learned from an unusually easy labeled sample may not
    # transfer to new wording. The cap is a deliberate deployment safety
    # buffer for high-recall triage, not a claim of probability calibration.
    if max_operating_threshold is not None:
        feasible = feasible[feasible["threshold"] <= max_operating_threshold + 1e-12]
    if feasible.empty:
        chosen = table.sort_values(
            ["recall", "min_fold_recall", "work_saved_fraction"], ascending=False
        ).iloc[0]
    else:
        chosen = feasible.sort_values(
            ["work_saved_fraction", "npv", "threshold"], ascending=False
        ).iloc[0]
    return float(chosen["threshold"]), table


def _oof_models(
    texts: Sequence[str],
    y: np.ndarray,
    groups: np.ndarray,
    config: RunConfig,
    progress: Callable[[str], None],
) -> Tuple[Dict[str, np.ndarray], np.ndarray]:
    splits = _folds(y, groups, config.cv_folds, config.random_seed)
    fold_ids = np.full(len(y), -1, dtype=int)
    probabilities = {
        "WORD_TFIDF": np.zeros(len(y), dtype=float),
        "CHAR_TFIDF": np.zeros(len(y), dtype=float),
        "CRITERIA": np.zeros(len(y), dtype=float),
    }
    for fold, (tr, va) in enumerate(splits):
        progress(f"교차검증 {fold + 1}/{len(splits)}")
        fold_ids[va] = fold
        train_text = [texts[i] for i in tr]
        valid_text = [texts[i] for i in va]
        for name, make_vec in (("WORD_TFIDF", _word_vectorizer), ("CHAR_TFIDF", _char_vectorizer)):
            vec = make_vec(len(tr))
            xtr = vec.fit_transform(train_text)
            xva = vec.transform(valid_text)
            clf = _classifier(config.random_seed + fold)
            clf.fit(xtr, y[tr])
            probabilities[name][va] = clf.predict_proba(xva)[:, 1]
        # Refit the criteria vectorizer on train-only text each fold (same
        # discipline as WORD/CHAR TF-IDF above). Fitting it once globally
        # would let validation-fold text leak into the IDF statistics used
        # to score that same fold, biasing OOF recall/precision optimistically.
        criteria_tr, criteria_va, _ = _criteria_matrix(train_text, valid_text, config)
        cclf = _classifier(config.random_seed + 100 + fold)
        cclf.fit(criteria_tr, y[tr])
        probabilities["CRITERIA"][va] = cclf.predict_proba(criteria_va)[:, 1]
    probabilities["ENSEMBLE"] = (
        0.45 * probabilities["WORD_TFIDF"]
        + 0.45 * probabilities["CHAR_TFIDF"]
        + 0.10 * probabilities["CRITERIA"]
    )
    return probabilities, fold_ids


def _model_summary(
    y: np.ndarray,
    probabilities: Dict[str, np.ndarray],
    fold_ids: np.ndarray,
    config: RunConfig,
) -> Tuple[pd.DataFrame, Dict[str, float], Dict[str, pd.DataFrame]]:
    rows, threshold_tables = [], {}
    for name, p in probabilities.items():
        threshold, table = choose_threshold(
            y, p, fold_ids, config.target_recall, max_operating_threshold=0.50
        )
        metrics = _metric_dict(y, p, threshold, fold_ids)
        metrics.update({
            "model": name,
            "roc_auc": float(roc_auc_score(y, p)),
            "average_precision": float(average_precision_score(y, p)),
            "brier_score": float(brier_score_loss(y, p)),
        })
        rows.append(metrics)
        threshold_tables[name] = table
    summary = pd.DataFrame(rows)
    eligible = summary[
        (summary["recall"] >= config.target_recall - 1e-12)
        & (summary["min_fold_recall"] >= max(0, config.target_recall - 0.05) - 1e-12)
    ]
    pool = eligible if not eligible.empty else summary
    champion_row = pool.sort_values(
        ["work_saved_fraction", "average_precision", "npv"], ascending=False
    ).iloc[0].to_dict()
    return summary, champion_row, threshold_tables


def _fit_full_models(
    labeled_texts: Sequence[str],
    unlabeled_texts: Sequence[str],
    y: np.ndarray,
    criteria_lab: np.ndarray,
    criteria_un: np.ndarray,
    config: RunConfig,
) -> Tuple[Dict[str, np.ndarray], Dict[str, object]]:
    probabilities: Dict[str, np.ndarray] = {}
    bundle: Dict[str, object] = {}
    for name, make_vec in (("WORD_TFIDF", _word_vectorizer), ("CHAR_TFIDF", _char_vectorizer)):
        vec = make_vec(len(y))
        xlab = vec.fit_transform(labeled_texts)
        xun = vec.transform(unlabeled_texts)
        clf = _classifier(config.random_seed)
        clf.fit(xlab, y)
        probabilities[name] = clf.predict_proba(xun)[:, 1] if len(unlabeled_texts) else np.array([])
        bundle[f"{name}_vectorizer"] = vec
        bundle[f"{name}_classifier"] = clf
    cclf = _classifier(config.random_seed + 100)
    cclf.fit(criteria_lab, y)
    probabilities["CRITERIA"] = cclf.predict_proba(criteria_un)[:, 1] if len(unlabeled_texts) else np.array([])
    bundle["CRITERIA_classifier"] = cclf
    probabilities["ENSEMBLE"] = (
        0.45 * probabilities["WORD_TFIDF"]
        + 0.45 * probabilities["CHAR_TFIDF"]
        + 0.10 * probabilities["CRITERIA"]
    )
    return probabilities, bundle


def _criteria_reason(
    feature_rows: np.ndarray,
    criteria_info: Dict[str, object],
) -> Tuple[List[str], List[float]]:
    names = list(criteria_info.get("feature_names", []))
    exclusion_idx = [i for i, x in enumerate(names) if x.startswith("Exclusion_")]
    criteria = dict(criteria_info.get("criteria", []))
    reasons, scores = [], []
    for row in feature_rows:
        if not exclusion_idx:
            reasons.append("")
            scores.append(0.0)
            continue
        local = row[exclusion_idx]
        best_local = int(np.argmax(local))
        idx = exclusion_idx[best_local]
        score = float(row[idx])
        name = names[idx]
        reasons.append(criteria.get(name, "") if score > 0 else "")
        scores.append(score)
    return reasons, scores


def _build_audit_sample(low: pd.DataFrame, n: int, seed: int) -> pd.DataFrame:
    if low.empty or n <= 0:
        return low.head(0).copy()
    n = min(n, len(low))
    boundary_n = min(math.ceil(n / 2), len(low))
    boundary = low.nlargest(boundary_n, "AI_Probability")
    remainder = low.drop(boundary.index)
    random_n = min(n - len(boundary), len(remainder))
    random_part = remainder.sample(random_n, random_state=seed) if random_n else remainder.head(0)
    audit = pd.concat([boundary, random_part]).drop_duplicates().head(n).copy()
    audit["Audit_Source"] = np.where(
        audit.index.isin(boundary.index), "Near safe-exclude boundary", "Random low-priority audit"
    )
    audit["Reviewer_Decision"] = ""
    audit["Exclusion_Reason"] = ""
    audit["Reviewer_Note"] = ""
    return audit


def _public_columns(frame: pd.DataFrame) -> pd.DataFrame:
    preferred = [
        "No.", "Year", "Title", "Abstract", "Human_Decision", "AI_Prediction",
        "Screening_Zone", "AI_Probability_pct", "Review_Threshold_pct",
        "Safe_Exclude_Threshold_pct", "Model", "Most_Similar_Exclusion_Criterion",
        "Exclusion_Criterion_Similarity", "Reviewer_Decision", "Exclusion_Reason",
        "Reviewer_Note", "Audit_Source", "OOF_Error_Type",
    ]
    columns = [x for x in preferred if x in frame.columns]
    extra = [x for x in frame.columns if x not in columns and not str(x).startswith("_")]
    return frame[columns + extra]


def _with_fresh_number(frame: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with one clean, sequential ``No.`` output column.

    Source workbooks commonly already contain a ``No.`` column. Preserve its
    first value column as ``Source_No.`` and replace ``No.`` instead of calling
    DataFrame.insert on top of the existing name.
    """
    out = frame.copy()
    if "No." in out.columns:
        existing = out.loc[:, "No."]
        if isinstance(existing, pd.DataFrame):
            existing = existing.iloc[:, 0]
        if "Source_No." not in out.columns:
            out["Source_No."] = existing.to_numpy()
        out = out.drop(columns=["No."])
    out.insert(0, "No.", np.arange(1, len(out) + 1))
    return out


def _format_workbook(path: Path, zone_sheets: Optional[Dict[str, str]] = None) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="173F5F")
    header_font = Font(color="FFFFFF", bold=True)
    fills = {
        ZONE_REVIEW: PatternFill("solid", fgColor="FFF2CC"),
        ZONE_MANUAL: PatternFill("solid", fgColor="DDEBF7"),
        ZONE_LOW: PatternFill("solid", fgColor="E2F0D9"),
        ZONE_HUMAN_INCLUDE: PatternFill("solid", fgColor="FFE699"),
        ZONE_HUMAN_EXCLUDE: PatternFill("solid", fgColor="E7E6E6"),
    }
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.sheet_view.showGridLines = False
        if ws.max_row >= 1:
            ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        headers = {cell.value: cell.column for cell in ws[1]}
        zone_col = headers.get("Screening_Zone")
        pct_cols = [idx for name, idx in headers.items() if name and (str(name).endswith("_pct") or "Similarity" in str(name))]
        for row in range(2, ws.max_row + 1):
            if zone_col:
                zone = ws.cell(row, zone_col).value
                if zone in fills:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row, col).fill = fills[zone]
            for col in pct_cols:
                ws.cell(row, col).number_format = "0.00"
            for col in range(1, ws.max_column + 1):
                ws.cell(row, col).alignment = Alignment(vertical="top", wrap_text=True)
        for col_idx, cell in enumerate(ws[1], start=1):
            name = str(cell.value or "")
            if name in {"Title"}:
                width = 58
            elif name == "제목":
                width = 100
            elif name in {"순번", "연도"}:
                width = 12
            elif name == "판정(o/x)":
                width = 16
            elif name in {"Abstract"}:
                width = 85
            elif "Criterion" in name or "Note" in name or "Reason" in name:
                width = 45
            else:
                values = [len(clean_text(ws.cell(r, col_idx).value)) for r in range(1, min(ws.max_row, 100) + 1)]
                width = min(max(max(values, default=8) + 2, 10), 28)
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 34
    wb.save(path)


def _save_dashboard(
    y: np.ndarray,
    probabilities: Dict[str, np.ndarray],
    champion: str,
    threshold: float,
    fold_ids: np.ndarray,
    path: Path,
) -> None:
    p = probabilities[champion]
    pred = (p >= threshold).astype(int)
    tn, fp, fn, tp = confusion_matrix(y, pred, labels=[0, 1]).ravel()
    fig, axes = plt.subplots(2, 3, figsize=(16, 9))
    ax = axes[0, 0]
    precision, recall, _ = precision_recall_curve(y, p)
    ax.plot(recall, precision, color="#1F77B4", lw=2)
    ax.set(xlabel="Recall", ylabel="Precision", title=f"Precision–Recall (AP={average_precision_score(y,p):.3f})")
    ax.grid(alpha=.25)
    ax = axes[0, 1]
    fpr, tpr, _ = roc_curve(y, p)
    ax.plot(fpr, tpr, color="#00A896", lw=2)
    ax.plot([0, 1], [0, 1], "--", color="gray")
    ax.set(xlabel="False-positive rate", ylabel="True-positive rate", title=f"ROC (AUC={roc_auc_score(y,p):.3f})")
    ax.grid(alpha=.25)
    ax = axes[0, 2]
    matrix = np.array([[tn, fp], [fn, tp]])
    ax.imshow(matrix, cmap="Blues")
    for i in range(2):
        for j in range(2):
            ax.text(j, i, str(matrix[i, j]), ha="center", va="center", fontsize=16)
    ax.set(xticks=[0,1], yticks=[0,1], xticklabels=["Exclude","Include"], yticklabels=["Exclude","Include"], xlabel="Predicted", ylabel="Human", title="OOF confusion matrix")
    ax = axes[1, 0]
    ax.hist(p[y == 0], bins=20, alpha=.65, label="Human exclude", color="#8E9AAF")
    ax.hist(p[y == 1], bins=20, alpha=.65, label="Human include", color="#F4A261")
    ax.axvline(threshold, color="#D00000", linestyle="--", label=f"Threshold {threshold:.3f}")
    ax.set(xlabel="OOF include probability", ylabel="Records", title="Probability distribution")
    ax.legend(fontsize=8)
    ax = axes[1, 1]
    names = list(probabilities)
    aps = [average_precision_score(y, probabilities[x]) for x in names]
    ax.barh(names, aps, color=["#457B9D", "#2A9D8F", "#E9C46A", "#F4A261"][:len(names)])
    ax.set_xlim(0, 1)
    ax.set(xlabel="Average precision", title="Model comparison")
    ax = axes[1, 2]
    folds = sorted(np.unique(fold_ids))
    recalls = [recall_score(y[fold_ids == f], pred[fold_ids == f], zero_division=0) for f in folds]
    ax.bar([str(x + 1) for x in folds], recalls, color="#6A4C93")
    ax.axhline(recall_score(y, pred), color="#D00000", linestyle="--", label="Overall")
    ax.set_ylim(0, 1.05)
    ax.set(xlabel="CV fold", ylabel="Recall", title="Fold stability")
    ax.legend(fontsize=8)
    fig.suptitle(f"SR Triage validation dashboard — {champion}", fontsize=16, fontweight="bold")
    fig.tight_layout(rect=(0, 0, 1, .96))
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _save_threshold_figure(table: pd.DataFrame, chosen: float, safe: float, path: Path) -> None:
    ordered = table.sort_values("threshold")
    fig, ax1 = plt.subplots(figsize=(11, 6))
    ax1.plot(ordered["threshold"], ordered["recall"], label="Recall", lw=2, color="#D62828")
    ax1.plot(ordered["threshold"], ordered["npv"], label="NPV", lw=2, color="#2A9D8F")
    ax1.set(xlabel="Include threshold", ylabel="Recall / NPV", ylim=(0, 1.03))
    ax2 = ax1.twinx()
    ax2.plot(ordered["threshold"], ordered["work_saved_fraction"], label="Work saved", lw=2, color="#457B9D")
    ax2.set(ylabel="Work saved fraction", ylim=(0, 1.03))
    ax1.axvline(chosen, color="#111111", linestyle="--", label=f"Review threshold {chosen:.3f}")
    ax1.axvline(safe, color="#F4A261", linestyle=":", lw=3, label=f"Safe-exclude boundary {safe:.3f}")
    lines = ax1.get_lines() + ax2.get_lines()
    ax1.legend(lines, [x.get_label() for x in lines], loc="lower left")
    ax1.grid(alpha=.2)
    ax1.set_title("Threshold trade-off (OOF validation)", fontweight="bold")
    fig.tight_layout()
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _readme_frame(config: RunConfig, status: str) -> pd.DataFrame:
    return pd.DataFrame([
        {"Item": "Purpose", "Value": "High-recall triage: prioritize records to review and identify low-priority exclusion candidates."},
        {"Item": "Important", "Value": "AI Exclude is not a final exclusion. Human review/audit is required; human decisions always take precedence."},
        {"Item": "Validation", "Value": "Performance uses out-of-fold predictions from existing human labels, not training-set accuracy."},
        {"Item": "Project", "Value": config.project_name},
        {"Item": "Target recall", "Value": config.target_recall},
        {"Item": "Safe-exclude target recall", "Value": config.safe_recall},
        {"Item": "Status", "Value": status},
        {"Item": "Decision entry", "Value": "Fill Reviewer_Decision with o/include/1 or x/exclude/0, then upload this workbook as an external label file."},
    ])


def _criteria_frame(config: RunConfig) -> pd.DataFrame:
    rows = [
        {"Section": "P", "Criteria": config.population},
        {"Section": "I", "Criteria": config.intervention},
        {"Section": "C", "Criteria": config.comparator},
        {"Section": "O", "Criteria": config.outcomes},
        {"Section": "Inclusion", "Criteria": config.inclusion_criteria},
    ]
    rows.extend({"Section": f"Exclusion {i+1}", "Criteria": x} for i, x in enumerate(config.exclusion_lines()))
    return pd.DataFrame(rows)


def create_seed_batch(
    records: pd.DataFrame,
    config: RunConfig,
    run_dir: Path,
    data_quality: pd.DataFrame,
    label_audit: pd.DataFrame,
    batch_size: int = 150,
) -> PipelineResult:
    usable = records[records["Title"] != ""].copy().reset_index(drop=True)
    unlabeled = usable[usable["Human_Label"].isna()].drop_duplicates(
        "_normalized_title", keep="first"
    ).copy()
    n = min(batch_size, len(unlabeled))
    if n == 0:
        raise ValueError("라벨이 없는 문헌이 없습니다. 라벨 열 매핑을 확인하세요.")
    texts = unlabeled["_text"].tolist()
    criteria_lab, _, _ = _criteria_matrix(texts, [], config)
    include_cols = len(config.criteria_blocks())
    if criteria_lab.size and include_cols:
        relevance = criteria_lab[:, :include_cols].max(axis=1)
    elif criteria_lab.size:
        # With exclusion criteria only, prioritize records least similar to an
        # exclusion rule so the starter batch is not dominated by easy negatives.
        relevance = 1.0 - criteria_lab.max(axis=1)
    else:
        relevance = np.zeros(len(unlabeled))
    chosen: List[int] = []
    top_n = min(math.ceil(n * .45), len(unlabeled))
    chosen.extend(np.argsort(-relevance)[:top_n].tolist())
    remaining = [i for i in range(len(unlabeled)) if i not in set(chosen)]
    diverse_n = min(math.ceil(n * .35), len(remaining))
    if diverse_n > 1 and len(remaining) > 2:
        vec = _word_vectorizer(len(remaining))
        x = vec.fit_transform([texts[i] for i in remaining])
        dims = min(30, x.shape[0] - 1, x.shape[1] - 1)
        dense = TruncatedSVD(n_components=max(1, dims), random_state=config.random_seed).fit_transform(x) if dims >= 1 else x.toarray()
        k = min(diverse_n, len(remaining))
        km = MiniBatchKMeans(n_clusters=k, random_state=config.random_seed, n_init=3, batch_size=256)
        labels = km.fit_predict(dense)
        for cluster in range(k):
            members = np.where(labels == cluster)[0]
            if len(members):
                center = km.cluster_centers_[cluster]
                local = members[np.argmin(np.linalg.norm(dense[members] - center, axis=1))]
                chosen.append(remaining[int(local)])
    remaining = [i for i in range(len(unlabeled)) if i not in set(chosen)]
    rng = np.random.default_rng(config.random_seed)
    if remaining and len(chosen) < n:
        chosen.extend(rng.choice(remaining, size=min(n - len(chosen), len(remaining)), replace=False).tolist())
    batch = unlabeled.iloc[chosen[:n]].copy().reset_index(drop=True)
    labeling_sheet = pd.DataFrame({
        "순번": np.arange(1, len(batch) + 1),
        "연도": batch["Year"].to_numpy(),
        "제목": batch["Title"].to_numpy(),
        "판정(o/x)": [""] * len(batch),
    })
    excel_path = run_dir / "1차_제목_라벨링용.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        labeling_sheet.to_excel(writer, sheet_name="1차_제목_라벨링", index=False)
        pd.DataFrame([
            {"안내": "판정(o/x) 열에 Include는 o, Exclude는 x를 입력하세요."},
            {"안내": "완료 후 같은 제목 원본과 함께 플랫폼의 '이전 라벨/검토 완료 파일'에 이 파일을 올리세요."},
            {"안내": "인간 판정이 충분해지면 모델 학습, 전체 예측, 성능평가가 자동 실행됩니다."},
        ]).to_excel(writer, sheet_name="사용방법", index=False)
    _format_workbook(excel_path)
    metadata_path = run_dir / "run_metadata.json"
    metadata_path.write_text(json.dumps({
        "status": "needs_labels",
        "config": asdict(config),
        "created_at": datetime.now().isoformat(),
        "starter_batch_n": len(batch),
    }, ensure_ascii=False, indent=2), encoding="utf-8")
    zip_path = run_dir / "1차_제목_라벨링_패키지.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.write(excel_path, excel_path.name)
        zf.write(metadata_path, metadata_path.name)
    counts = {
        "total": len(records),
        "human_include": int((records["Human_Label"] == 1).sum()),
        "human_exclude": int((records["Human_Label"] == 0).sum()),
        "starter_batch": len(batch),
    }
    return PipelineResult(
        status="needs_labels", run_dir=run_dir, excel_path=excel_path, zip_path=zip_path,
        dashboard_path=None, threshold_path=None, model_path=None, metrics={}, counts=counts,
        message=f"성능평가에 필요한 인간 라벨이 부족하여 {len(batch)}편의 시작 라벨링 배치를 생성했습니다.",
    )


def run_pipeline(
    records: pd.DataFrame,
    config: RunConfig,
    output_root: Path,
    label_audit: Optional[pd.DataFrame] = None,
    progress: Optional[Callable[[str], None]] = None,
) -> PipelineResult:
    progress = progress or (lambda _message: None)
    label_audit = label_audit if label_audit is not None else pd.DataFrame()
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_name = re.sub(r"[^a-zA-Z0-9가-힣._-]+", "_", config.project_name).strip("_") or "SR_Project"
    run_dir = output_root / safe_name / "runs" / stamp
    run_dir.mkdir(parents=True, exist_ok=True)
    data_quality = validate_records(records)
    valid = records[records["Title"] != ""].copy().reset_index(drop=True)
    labeled_for_checks = valid[valid["Human_Label"].notna()].copy()
    label_conflicts = labeled_for_checks.groupby("_normalized_title")["Human_Label"].nunique()
    conflicting_titles = label_conflicts[label_conflicts > 1]
    if len(conflicting_titles):
        examples = ", ".join(list(conflicting_titles.index[:3]))
        raise ValueError(
            f"같은 제목에 서로 다른 인간 판정이 {len(conflicting_titles)}건 있습니다. "
            f"라벨 충돌을 먼저 수정하세요. 예: {examples}"
        )
    unique_labeled = labeled_for_checks.drop_duplicates("_normalized_title")
    unique_includes = int((unique_labeled["Human_Label"] == 1).sum())
    unique_excludes = int((unique_labeled["Human_Label"] == 0).sum())
    includes = int((valid["Human_Label"] == 1).sum())
    excludes = int((valid["Human_Label"] == 0).sum())
    if min(unique_includes, unique_excludes) < config.min_per_class:
        return create_seed_batch(valid, config, run_dir, data_quality, label_audit)

    progress("인간 라벨과 예측 대상을 분리하는 중")
    labeled = valid[valid["Human_Label"].notna()].copy().reset_index(drop=True)
    unlabeled = valid[valid["Human_Label"].isna()].copy().reset_index(drop=True)
    y = labeled["Human_Label"].astype(int).to_numpy()
    groups = labeled["_group"].astype(str).to_numpy()
    lab_text = labeled["_text"].tolist()
    un_text = unlabeled["_text"].tolist()
    criteria_lab, criteria_un, criteria_info = _criteria_matrix(lab_text, un_text, config)

    progress("Word/Character TF-IDF 고재현율 모델을 검증하는 중")
    oof, fold_ids = _oof_models(lab_text, y, groups, config, progress)
    summary, champion_row, threshold_tables = _model_summary(y, oof, fold_ids, config)
    champion = str(champion_row["model"])
    review_threshold = float(champion_row["threshold"])
    safe_threshold, safe_table = choose_threshold(
        y,
        oof[champion],
        fold_ids,
        max(config.safe_recall, config.target_recall),
        fold_tolerance=.02,
        max_operating_threshold=min(0.25, review_threshold * 0.50),
    )

    progress("전체 문헌에 모델을 적용하는 중")
    un_probs, model_bundle = _fit_full_models(
        lab_text, un_text, y, criteria_lab, criteria_un, config
    )
    champion_un = un_probs[champion]
    reasons, reason_scores = _criteria_reason(criteria_un, criteria_info)
    unlabeled["AI_Probability"] = champion_un
    unlabeled["AI_Probability_pct"] = champion_un * 100
    unlabeled["Review_Threshold_pct"] = review_threshold * 100
    unlabeled["Safe_Exclude_Threshold_pct"] = safe_threshold * 100
    unlabeled["Model"] = champion
    unlabeled["AI_Prediction"] = np.where(champion_un >= review_threshold, "Include", "Exclude")
    unlabeled["Screening_Zone"] = np.select(
        [champion_un >= review_threshold, champion_un < safe_threshold],
        [ZONE_REVIEW, ZONE_LOW],
        default=ZONE_MANUAL,
    )
    unlabeled["Most_Similar_Exclusion_Criterion"] = reasons
    unlabeled["Exclusion_Criterion_Similarity"] = np.asarray(reason_scores) * 100
    unlabeled["Human_Decision"] = ""
    unlabeled["Reviewer_Decision"] = ""
    unlabeled["Exclusion_Reason"] = ""
    unlabeled["Reviewer_Note"] = ""

    labeled_eval = labeled.copy()
    labeled_eval["AI_Probability"] = oof[champion]
    labeled_eval["AI_Probability_pct"] = oof[champion] * 100
    labeled_eval["Review_Threshold_pct"] = review_threshold * 100
    labeled_eval["Safe_Exclude_Threshold_pct"] = safe_threshold * 100
    labeled_eval["Model"] = champion
    labeled_eval["Human_Decision"] = np.where(y == 1, "Include", "Exclude")
    labeled_eval["AI_Prediction"] = np.where(oof[champion] >= review_threshold, "Include", "Exclude")
    labeled_eval["Screening_Zone"] = np.where(y == 1, ZONE_HUMAN_INCLUDE, ZONE_HUMAN_EXCLUDE)
    labeled_eval["OOF_Error_Type"] = np.select(
        [(y == 1) & (oof[champion] < review_threshold), (y == 0) & (oof[champion] >= review_threshold)],
        ["FALSE_NEGATIVE_REVIEW_FIRST", "FALSE_POSITIVE"], default="CORRECT",
    )

    review_first = unlabeled[unlabeled["Screening_Zone"] == ZONE_REVIEW].sort_values("AI_Probability", ascending=False)
    manual = unlabeled[unlabeled["Screening_Zone"] == ZONE_MANUAL].sort_values("AI_Probability", ascending=False)
    low = unlabeled[unlabeled["Screening_Zone"] == ZONE_LOW].sort_values("AI_Probability", ascending=False)
    audit = _build_audit_sample(low, config.audit_size, config.random_seed)
    false_negatives = labeled_eval[
        labeled_eval["OOF_Error_Type"] == "FALSE_NEGATIVE_REVIEW_FIRST"
    ].sort_values("AI_Probability")

    all_records = _with_fresh_number(
        pd.concat([labeled_eval, unlabeled], ignore_index=True, sort=False)
    )
    review_first = _with_fresh_number(review_first)
    manual = _with_fresh_number(manual)
    low = _with_fresh_number(low)
    audit = _with_fresh_number(audit)
    labeled_eval = _with_fresh_number(labeled_eval)
    false_negatives = _with_fresh_number(false_negatives)

    progress("Excel과 성능 피규어를 생성하는 중")
    dashboard_path = run_dir / "SR_Triage_PERFORMANCE_DASHBOARD.png"
    threshold_path = run_dir / "SR_Triage_THRESHOLD_TRADEOFF.png"
    _save_dashboard(y, oof, champion, review_threshold, fold_ids, dashboard_path)
    _save_threshold_figure(threshold_tables[champion], review_threshold, safe_threshold, threshold_path)

    metrics = dict(champion_row)
    metrics.update({
        "champion": champion,
        "review_threshold": review_threshold,
        "safe_exclude_threshold": safe_threshold,
        "safe_target_recall": max(config.safe_recall, config.target_recall),
        "labeled_n": len(labeled),
    })
    counts = {
        "total": len(valid),
        "human_include": includes,
        "human_exclude": excludes,
        "unlabeled": len(unlabeled),
        "review_first": len(review_first),
        "manual_review": len(manual),
        "low_priority": len(low),
        "audit_sample": len(audit),
    }
    flow = pd.DataFrame([counts])
    excel_path = run_dir / "SR_Triage_FINAL_RESULTS.xlsx"
    with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
        _readme_frame(config, "Model trained and OOF validated").to_excel(writer, sheet_name="00_README", index=False)
        _public_columns(all_records).to_excel(writer, sheet_name="ALL_RECORDS", index=False)
        _public_columns(review_first).to_excel(writer, sheet_name="01_REVIEW_FIRST", index=False)
        _public_columns(manual).to_excel(writer, sheet_name="02_MANUAL_REVIEW", index=False)
        _public_columns(low).to_excel(writer, sheet_name="03_LOW_PRIORITY", index=False)
        _public_columns(audit).to_excel(writer, sheet_name="04_LOW_PRIORITY_AUDIT", index=False)
        _public_columns(labeled_eval).to_excel(writer, sheet_name="LABELED_OOF", index=False)
        _public_columns(false_negatives).to_excel(writer, sheet_name="FALSE_NEGATIVES", index=False)
        summary.to_excel(writer, sheet_name="MODEL_PERFORMANCE", index=False)
        threshold_tables[champion].to_excel(writer, sheet_name="THRESHOLD_CURVE", index=False)
        safe_table.to_excel(writer, sheet_name="SAFE_THRESHOLD_CURVE", index=False)
        flow.to_excel(writer, sheet_name="FLOW_SUMMARY", index=False)
        _criteria_frame(config).to_excel(writer, sheet_name="CRITERIA", index=False)
        data_quality.to_excel(writer, sheet_name="DATA_QUALITY", index=False)
        label_audit.to_excel(writer, sheet_name="LABEL_AUDIT", index=False)
    _format_workbook(excel_path)
    wb = load_workbook(excel_path)
    ws = wb.create_sheet("FIGURES")
    ws.sheet_view.showGridLines = False
    ws.add_image(XLImage(str(dashboard_path)), "A1")
    ws.add_image(XLImage(str(threshold_path)), "A48")
    wb.save(excel_path)

    model_bundle.update({
        "config": asdict(config),
        "champion": champion,
        "review_threshold": review_threshold,
        "safe_exclude_threshold": safe_threshold,
        "criteria_info": criteria_info,
        "model_weights": {"WORD_TFIDF": .45, "CHAR_TFIDF": .45, "CRITERIA": .10},
    })
    model_path = run_dir / "SR_Triage_MODEL.joblib"
    joblib.dump(model_bundle, model_path, compress=3)
    metadata_path = run_dir / "run_metadata.json"
    metadata_path.write_text(json.dumps({
        "status": "completed", "created_at": datetime.now().isoformat(),
        "config": asdict(config), "metrics": metrics, "counts": counts,
        "method_note": "Out-of-fold validation; AI exclude is a low-priority recommendation, not a final decision.",
    }, ensure_ascii=False, indent=2, default=float), encoding="utf-8")
    zip_path = run_dir / "SR_Triage_FINAL_PACKAGE.zip"
    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for path in (excel_path, dashboard_path, threshold_path, model_path, metadata_path):
            zf.write(path, path.name)
    return PipelineResult(
        status="completed", run_dir=run_dir, excel_path=excel_path, zip_path=zip_path,
        dashboard_path=dashboard_path, threshold_path=threshold_path, model_path=model_path,
        metrics=metrics, counts=counts,
        message="고재현율 교차검증, 전체 예측, 저우선순위 감사표본, Excel 및 피규어 생성이 완료되었습니다.",
    )
